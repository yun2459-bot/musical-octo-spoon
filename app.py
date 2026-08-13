# -*- coding: utf-8 -*-
"""
AI 기반 현장 안전 데이터 통합 분석 대시보드 (프로토타입)
- 구성: ① 전사 현황  ② 지사 상세  (2개 탭, 안전관리자 친화적 단순화)
실행:  streamlit run app.py
"""
import base64
import html
import io
import json
import os
import re
import subprocess
import sys
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm
import plotly.express as px
import plotly.graph_objects as go
from wordcloud import WordCloud

import scoring as S
import heatwave as HW

# 온열질환·조치 현황 구글폼(GOOGLE_SHEET_CSV_URL이 읽는 그 폼)의 "지사" 문항을 URL
# 파라미터로 미리 채워 여는 링크 — 구글폼 자체의 "URL 미리 채우기" 기능이라 폼이 안
# 바뀌는 한 별도 유지보수가 필요 없다.
_FORM_BASE_URL = "https://docs.google.com/forms/d/e/1FAIpQLSe8QFLQbs3KwCUVMqHFv-C0gLVdDcILhcsctKUDHbHP8DgGRg/viewform"

# 현장 사진 전용 구글폼(별도 폼, PHOTO_SHEET_CSV_URL이 읽는 그 폼) 제출 링크. 파일 업로드
# 문항이 있는 폼이라 구글 정책상 응답자가 구글 계정으로 로그인해야 제출 가능하다(폼 설정으로
# 끌 수 없는 제약 — 사내 구글 계정으로 로그인하면 정상 제출됨).
_PHOTO_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSd4IeOzIzwZTfTuEv5JT0nr8eDBNrNZMs_EwXRisyayhhj66Q/viewform"

# 작업중지 즉시 보고 전용 구글폼(별도 폼, HW.STOPPAGE_SHEET_CSV_URL이 읽는 그 폼) 제출 링크.
_STOPPAGE_FORM_URL = "https://docs.google.com/forms/d/1pk8mc8krfbTuocYAO0MZW1peD5z8K288aRjl1XKXR6c/viewform"

# 온열질환 발생 즉시 보고 전용 구글폼(별도 폼, HW.PATIENT_SHEET_CSV_URL이 읽는 그 폼) 제출 링크.
_PATIENT_FORM_URL = "https://docs.google.com/forms/d/1hJrpkQfOviMaUNnb5bRvyRm-5XwERq1fPeoSfAdJCN4/viewform"


def _branch_form_url(branch: str) -> str:
    return f"{_FORM_BASE_URL}?usp=pp_url&entry.996594318={quote(branch)}"


def _bulk_photo_template_bytes() -> bytes:
    """엑셀 일괄 사진 등록용 빈 양식 — "사진목록" 시트(지사/활동내용/사진파일명)와
    "필독사항" 시트(작성 규칙 + 정확한 지사명 철자 참고용) 두 장으로 구성한다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "사진목록"
    ws1.append(["지사", "활동내용", "사진파일명"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    # 예전엔 이 아래에 "예: 광양 / 예: ... / 예: photo1.jpg" 예시 행을 넣어뒀는데,
    # 실제 데이터인 줄 알고 일부 칸만 고쳐 쓰다 "예: photo1.jpg"가 그대로 남아
    # 매칭이 실패하는 사고가 있었다(2026-08-06) — 예시는 아래 "필독사항" 시트에
    # 텍스트로만 보여주고, 이 시트엔 실제로 채워 넣을 빈 칸만 남긴다.
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 22

    ws2 = wb.create_sheet("필독사항")
    ws2.cell(row=1, column=1, value="★ 작성 전 꼭 확인해주세요 ★").font = Font(bold=True, color="C81D25", size=13)
    notes = [
        "1. '사진파일명' 칸은 실제로 함께 첨부하는 사진 파일의 이름과 글자 하나까지 정확히 같아야 "
        "합니다(대소문자·확장자·띄어쓰기 포함). 다르면 그 행은 등록되지 않습니다.",
        "2. '지사' 칸은 아래 지사 목록의 철자와 정확히 같아야 합니다(띄어쓰기도 동일해야 함).",
        "3. 한 행 = 사진 1장입니다. 여러 장을 올리려면 행을 여러 개로 나눠 적어주세요"
        "(같은 지사를 여러 행에 반복해서 적어도 됩니다).",
        "4. 사진 파일 형식은 jpg, jpeg, png만 가능합니다.",
        "5. 작성 예시 — 지사: 광양 / 활동내용: 그늘막 설치, 이온음료 배부 / "
        "사진파일명: photo1.jpg (실제 올리는 사진 파일 이름 그대로).",
        "6. 등록 버튼을 누른 뒤 화면(사진 목록·보고서)에 반영되기까지 최대 1분 정도 걸릴 수 있습니다.",
    ]
    row = 3
    for note in notes:
        ws2.cell(row=row, column=1, value=note)
        row += 1
    row += 1
    ws2.cell(row=row, column=1, value="지사 목록 (정확한 철자 참고용)").font = Font(bold=True)
    row += 1
    for branch in HW.branch_order():
        ws2.cell(row=row, column=1, value=branch)
        row += 1
    ws2.column_dimensions["A"].width = 90
    for r in range(3, 3 + len(notes)):
        ws2.cell(row=r, column=1).alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _print_report_html(week_start: pd.Timestamp | None = None) -> str:
    """주간 CSO·임원진 보고용 인쇄 리포트 본문 HTML.

    2페이지 구성 — 1p: 전체 현황 / 옥외 작업 중지 지사·상세 이유 / 주간 예방 점검
    결과(지사별), 2p: 지사별 활동 사진 1장씩. CSS의 page-break로 페이지를 나눈다.

    week_start(그 주 월요일)를 주면 그 주(과거 포함) 기준으로 다시 뽑고, 안 주면
    이번주 기준(기본 동작)이다 — CSO/대표가 지난주 보고서를 다시 요청할 때를
    대비해 모든 데이터를 주차 단위로 다시 계산할 수 있게 해둔 것.
    """
    if week_start is None:
        _, week_start = HW.this_and_last_week()

    esum = HW.patient_event_summary(week_start)
    obs = HW.load_observations()
    week_top = HW.week_top_temp_branch(obs, week_start)

    stoppages = HW.weekly_stoppage_reports(week_start)
    checklist = HW.weekly_checklist(week_start)
    # 보고서에 넣을 사진은 기본적으로 지사별 그 주 최신 1장이지만, 화면에서 직접 고른
    # 사진이 있으면 그것을 우선한다(st.session_state["_report_photo_pick"]: {지사: photo_url}).
    # 다른 주 사진은 weekly_photo_reports()가 걸러내므로 섞여 들어가지 않는다. 사진
    # 직접 고르기는 화면에 보이는 이번주 캐러셀 기준이라, 과거 주차 재출력 시에는
    # (그 주에 직접 고른 사진이 세션에 남아있을 리 없으므로) 자연히 그 주 최신 사진이
    # 쓰인다.
    photos = HW.weekly_photo_reports(week_start)
    picks = st.session_state.get("_report_photo_pick", {})
    if photos.empty:
        photo_by_branch = photos
    else:
        latest = photos.sort_values("timestamp", ascending=False).drop_duplicates(subset="branch")
        chosen = photos[photos.apply(
            lambda r: picks.get(r["branch"]) == r["photo_url"], axis=1)]
        # 고른 지사는 chosen 행으로, 나머지는 최신 행으로 채운다.
        photo_by_branch = pd.concat(
            [chosen, latest[~latest["branch"].isin(chosen["branch"])]], ignore_index=True)
    week_end = week_start + pd.Timedelta(days=4)  # 월~금 근무 주간 기준
    period_label = f"{week_start.month}월 {week_start.day}일 ~ {week_end.month}월 {week_end.day}일"

    top_line = (f"{week_top['branch']} · {week_top['apparent_temp']:.1f}℃" if week_top is not None
                else "관측 데이터 없음")
    top_alert = HW.weekly_top_alert_branch(week_start)
    top_alert_line = (f"{top_alert['branch']} · {HW.level_label(top_alert['level'])} 단계 "
                       f"({top_alert['count']}회)" if top_alert else "해당 주 발령 없음")
    section1 = f"""
    <table class="rpt-table">
      <tr><th>누적 온열질환 발생 인원</th><td>{esum['cumulative']}명</td></tr>
      <tr><th>해당 주 온열질환 발생</th><td>{esum['this_week']}명</td></tr>
      <tr><th>주간 최고 온도 사업장</th><td>{html.escape(top_line)}</td></tr>
      <tr><th>해당 주 최고 폭염특보 단계 지사</th><td>{html.escape(top_alert_line)}</td></tr>
      <tr><th>해당 주 옥외 작업 중지</th><td>{len(stoppages)}건</td></tr>
    </table>
    """

    stop_only = stoppages[stoppages["구분"].str.contains("중지", na=False)] if not stoppages.empty else stoppages
    if stop_only.empty:
        section2 = '<p class="rpt-empty">해당 주 작업중지 보고 없음</p>'
    else:
        rows2 = "".join(
            f"<tr><td>{html.escape(str(r.branch))}</td>"
            f"<td>{r.timestamp.strftime('%m-%d %H:%M')}</td>"
            f"<td>{html.escape(str(r.상세) or '(상세 미기재)')}</td></tr>"
            for r in stop_only.itertuples()
        )
        section2 = f'<table class="rpt-table"><tr><th>지사</th><th>제출시각</th><th>상세 이유</th></tr>{rows2}</table>'

    events = HW.weekly_patient_events(week_start)
    if events.empty:
        section_pt = '<p class="rpt-empty">해당 주 온열질환 발생 보고 없음</p>'
    else:
        rows_pt = "".join(
            f"<tr><td>{html.escape(str(r.branch))}</td>"
            f"<td>{r.timestamp.strftime('%m-%d %H:%M')}</td>"
            f"<td>{int(r.환자수)}명</td>"
            f'<td class="{"issue" if "열사병" in str(r.증상) else ""}">'
            f"{html.escape(str(r.증상) or '-')}</td>"
            f"<td>{html.escape(str(r.상세) or '(상세 미기재)')}</td></tr>"
            for r in events.itertuples()
        )
        section_pt = ('<table class="rpt-table">'
                      '<tr><th>지사</th><th>발생시각</th><th>환자 수</th><th>증상 구분</th>'
                      '<th>발생 경위 · 조치</th></tr>'
                      f'{rows_pt}</table>')
        # 주간 자기보고와 숫자가 어긋나는 지사는 보고서에도 남겨, 회의 자리에서 바로
        # 누락 지사를 확인할 수 있게 한다.
        mis = HW.patient_report_mismatch(week_start)
        if not mis.empty:
            _t = ", ".join(f"{r.branch}(즉시 {r.즉시보고}명 / 주간 {r.주간보고}명)"
                           for r in mis.itertuples())
            section_pt += (f'<p class="rpt-note">※ 즉시보고와 주간보고 인원이 다른 지사: '
                           f'{html.escape(_t)} — 누락 여부 확인 필요</p>')

    if checklist.empty:
        section3 = '<p class="rpt-empty">점검 데이터 없음</p>'
    else:
        head_cols = "".join(f"<th>{v}</th>" for v in HW.CHECKLIST_LABELS.values())
        rows3 = []
        def _cell_class(field: str, value: str) -> str:
            if "특이사항 기재" in value:
                return "issue"
            if not value.strip():
                return "no-submit"
            if value.strip() == HW.CHECKLIST_OK[field]:
                return "ok"
            return ""

        n_status = len(HW.CHECKLIST_FIELDS)
        for i, r in enumerate(checklist.itertuples()):
            # 상태행 + (있으면) 특이사항행을 "지사 한 묶음"으로 보이게 지사마다 배경을
            # 교대로 옅게 칠한다 — 묶음 구분이 없으니 특이사항행이 다음 지사 상태행과
            # 이어 붙어 보여 복잡해 보인다는 피드백 반영.
            grp = "grp-odd" if i % 2 else "grp-even"
            cells = "".join(
                f'<td class="{_cell_class(f, str(getattr(r, f)))}">'
                f'{html.escape(str(getattr(r, f)) or "-")}</td>'
                for f in HW.CHECKLIST_FIELDS
            )
            rows3.append(f'<tr class="{grp}"><td>{html.escape(str(r.branch))}</td>{cells}</tr>')
            # 구글폼 자유서술 특이사항은 줄바꿈으로 구조화해 적는 경우가 많고(접수일자/
            # 신고내용/조치결과 등 항목별 줄) 길이도 제각각이라, 좁은 옆 칸에 욱여넣는
            # 대신 내용이 있는 지사만 그 아래 폭 전체를 쓰는 병합 행을 따로 붙인다
            # (내용 없는 지사는 이 행 자체가 안 생겨 표가 불필요하게 늘어지지 않는다).
            # html.escape는 줄바꿈을 안 지워주지만 브라우저는 기본적으로 \n을 공백으로
            # 뭉개버리므로, <br>로 바꿔 원래 줄바꿈 구조를 인쇄물에도 그대로 살린다.
            note = html.escape(str(r.점검특이사항) or "").replace("\n", "<br>")
            if note.strip():
                rows3.append(f'<tr class="{grp} rpt-note-row"><td class="rpt-note-label">특이사항</td>'
                             f'<td colspan="{n_status}">{note}</td></tr>')
        # 지사 칸은 좁게, 상태 5칸은 짧은 라벨만 담으니 균등하게 나머지 폭을 쓴다.
        status_w = 88 / n_status
        colgroup = (
            '<colgroup><col style="width:12%">'
            + f'<col style="width:{status_w:.1f}%">' * n_status
            + '</colgroup>'
        )
        section3 = (f'<table class="rpt-table small">{colgroup}'
                    f'<tr><th>지사</th>{head_cols}</tr>{"".join(rows3)}</table>')

    # 사진이 없는 지사도 "그 지사엔 아직 없다"는 게 보이도록 빈 칸으로 채워서 전체
    # 지사를 다 늘어놓는다 — 있는 지사만 나오면 빠진 지사가 안 보여 헷갈린다.
    photo_map = ({p.branch: p for p in photo_by_branch.itertuples()} if not photo_by_branch.empty else {})

    def _photo_card(b: str) -> str:
        if b in photo_map:
            return (f'<div class="rpt-photo">'
                    f'<img src="{html.escape(photo_map[b].photo_url)}" />'
                    f'<div class="cap">{html.escape(b)} · '
                    f'{photo_map[b].timestamp.strftime("%m-%d %H:%M")}</div></div>')
        return (f'<div class="rpt-photo rpt-photo-empty">'
                f'<div class="rpt-photo-placeholder">사진 없음</div>'
                f'<div class="cap">{html.escape(b)}</div></div>')

    # 지사가 정확히 12곳이라 4열 x 3행 그리드 한 장(가로 방향)에 전부 들어간다.
    # 세로(portrait)인 1~4번 절과 달리 이 페이지만 가로(landscape)로 인쇄해
    # 칸 하나하나를 더 넓게 쓴다.
    _all = HW.branch_order()
    photo_pages = (
        f'<div class="rpt-page rpt-photo-page">'
        f'<h2>5. 온열질환 예방 활동사진</h2>'
        f'<div class="rpt-photo-grid">{"".join(_photo_card(b) for b in _all)}</div>'
        f'</div>'
    )

    return f"""
    <div class="rpt-page" id="page1">
      <h1>주간 온열질환 예방 대응 보고서</h1>
      <div class="rpt-date">기준기간: {period_label}</div>
      <h2>1. 전체 현황</h2>
      {section1}
      <h2>2. 옥외 작업 중지 지사 및 상세 이유</h2>
      {section2}
      <h2>3. 온열질환 발생 보고 상세</h2>
      {section_pt}
      <h2>4. 주간 온열질환 예방 확인 결과 (지사별)</h2>
      {section3}
    </div>
    {photo_pages}
    """


def _render_print_report_button(week_start: pd.Timestamp | None = None) -> None:
    """새 창을 열어 인쇄 전용 HTML을 그린 뒤 브라우저 인쇄 대화상자를 띄우는 버튼.

    components.html은 iframe 안에서 렌더링되므로 그 안에서 window.print()를 부르면
    iframe 자신만 인쇄하려 들어 브라우저마다 동작이 들쭉날쭉하다 — 대신 새 창을 열어
    독립된 문서로 print()를 호출하면 항상 그 문서 하나만 인쇄된다.

    week_start(그 주 월요일)를 주면 그 주 보고서를, 안 주면 이번주 보고서를 뽑는다.
    """
    css = """
    body { font-family: 'Malgun Gothic', 'SEBANG Gothic', sans-serif; color: #111; margin: 20px; }
    h1 { font-size: 20px; margin: 0 0 4px; }
    .rpt-date { color: #666; font-size: 12px; margin-bottom: 8px; }
    h2 { font-size: 15px; margin: 10px 0 4px; border-bottom: 2px solid #333; padding-bottom: 3px; }
    .rpt-table { width: 100%; border-collapse: collapse; font-size: 12px; margin-bottom: 6px; }
    /* 1~3번 절 표는 한 줄짜리 짧은 값만 담는데 줄 높이가 커서 세로 공간을 많이
       먹었다 — 한 장 맞춤 축소량을 줄이려 여백을 최소로 좁힘(2026-08-07). */
    .rpt-table th, .rpt-table td { border: 1px solid #ccc; padding: 2px 7px; text-align: left;
        line-height: 1.35; }
    .rpt-table th { background: #eee; }
    .rpt-table.small { table-layout: fixed; }
    .rpt-table.small th, .rpt-table.small td { font-size: 10.5px; padding: 3px 5px;
        word-break: keep-all; overflow-wrap: break-word; vertical-align: top; line-height: 1.5; }
    .rpt-table td.issue { background: #fdd; color: #900; font-weight: 700; }
    .rpt-table td.ok { background: #d9f2e3; color: #14532d; }
    /* 그 주에 아예 점검을 제출하지 않은 지사 — 특이사항(issue)과 구분해 따로 두되
       똑같이 눈에 띄는 빨강으로 칠해 미제출도 확인이 필요한 상태임을 드러낸다. */
    .rpt-table td.no-submit { background: #fdd; color: #900; }
    /* 특이사항 병합 행 — 위 상태 행과 한 묶음으로 보이도록 위쪽 테두리를 없앤다.
       배경은 지사 묶음 단위 줄무늬(grp-odd/even)를 그대로 물려받게 여기선 따로
       칠하지 않는다 — td에 직접 배경을 주면 아래 줄무늬 규칙보다 우선해버려서
       특이사항 행만 항상 같은 색이 되어(어느 지사든 구분이 안 됨) 묶음 구분 효과가
       사라진다. */
    .rpt-table.small tr.rpt-note-row td { border-top: none; }
    .rpt-table.small td.rpt-note-label { font-weight: 700; color: #555; white-space: nowrap; }
    /* 지사 한 묶음(상태행 + 있으면 특이사항행)마다 배경을 교대로 옅게 칠해 어디부터
       어디까지가 한 지사인지 한눈에 보이게 한다. td에 직접 배경(.ok/.issue)이 있는
       칸은 그 색이 그대로 우선되고, 나머지 칸만 이 줄무늬가 비쳐 보인다. */
    .rpt-table.small tr.grp-odd { background: #f2f2f2; }
    .rpt-empty { color: #888; font-size: 12px; }
    .rpt-note { color: #900; font-size: 11px; margin: 4px 0 0; }
    .rpt-page { page-break-after: always; }
    .rpt-page:last-child { page-break-after: auto; }
    /* 지사가 정확히 12곳이라 4열x3행 그리드로 한 면에 전부 들어간다. 이 페이지만
       가로(landscape)로 인쇄해 세로 페이지(1~4번 절)보다 칸을 더 넓게 쓴다. */
    @page landscape { size: A4 landscape; margin: 10mm; }
    .rpt-photo-page { page: landscape; }
    .rpt-photo-page h2 { margin-top: 0; }
    .rpt-photo-grid { display: grid; grid-template-columns: repeat(4, 1fr); grid-template-rows: repeat(3, 1fr);
        gap: 5mm; height: 168mm; }
    .rpt-photo { display: flex; flex-direction: column; min-height: 0; overflow: hidden; }
    .rpt-photo img { width: 100%; height: 100%; flex: 1 1 auto; min-height: 0; object-fit: cover;
        border-radius: 6px; }
    .rpt-photo .cap { font-size: 11px; color: #333; margin-top: 4px; flex: 0 0 auto; }
    .rpt-photo-placeholder { width: 100%; flex: 1 1 auto; min-height: 0; border-radius: 6px;
        background: #f2f2f2; border: 1px dashed #bbb; display: flex; align-items: center;
        justify-content: center; color: #999; font-size: 12px; }
    @page { size: A4; margin: 14mm; }
    /* @page 여백에 더해 body 여백까지 들어가면 실제 인쇄 영역이 182x269mm보다 작아져
       아래 JS의 "한 장 맞춤" 계산이 어긋난다 — 인쇄 때는 body 여백을 없앤다. */
    @media print { body { margin: 0; } }
    """
    full_html = f"<html><head><meta charset='utf-8'><title>주간 온열질환 예방 대응 보고서</title>" \
                f"<style>{css}</style></head><body>{_print_report_html(week_start)}</body></html>"
    doc_js = json.dumps(full_html).replace("</", "<\\/")
    button_html = f"""
    <button id="printReportBtn" style="padding:10px 18px; font-size:14px; font-weight:700;
        border:none; border-radius:8px; background:{SEBANG_ORANGE}; color:white; cursor:pointer; width:100%;">
      🖨️ 주간 보고서 인쇄
    </button>
    <script>
    document.getElementById('printReportBtn').onclick = function () {{
        var w = window.open('', '_blank');
        w.document.open();
        w.document.write({doc_js});
        w.document.close();
        var fitAndPrint = function () {{
            // 1~4번 절(id="page1")을 A4 한 장에 무조건 맞춘다.
            //
            // 예전엔 팝업 창(=임의의 화면 폭)에서 잰 높이로 축소 배율을 계산했는데,
            // 정작 인쇄는 A4 인쇄 폭(182mm)으로 "다시 배치"된다 — 폭이 좁아지면 표
            // 칸 안에서 줄바꿈이 훨씬 늘어나 실제 인쇄 높이가 화면 높이보다 크게
            // 커진다. 그래서 계산된 축소량이 애초에 부족했고, 최저 한도(0.55)를
            // 없애도 계속 2페이지로 넘어갔다(2026-08-07 원인 규명).
            // → 재기 전에 page1의 폭을 실제 인쇄 폭으로 고정해 화면 배치를 인쇄
            //   배치와 일치시킨다.
            //
            // 폭은 인쇄 폭에 "고정"하고 균일 축소만 한다. 한때 축소한 만큼 폭을
            // 1/배율로 넓혀 지면을 꽉 채우게 했는데, 그러면 요소 폭이 인쇄 폭보다
            // 넓어져 인쇄 시 배치가 어긋났다(2026-08-07 사용자 확인) — 폭을 넘기지
            // 않으면 최악의 경우에도 "덜 줄어들 뿐" 가로로 깨지지는 않는다.
            // 폭이 고정이면 zoom은 순수 배율이라 한 번 계산으로 정확히 떨어진다.
            var doc = w.document;
            var page = doc.getElementById('page1');
            if (page) {{
                // mm->px 환산은 브라우저·배율마다 달라 하드코딩하지 않고 직접 잰다.
                // 182x269mm = A4(210x297)에서 @page 여백 14mm를 뺀 실제 인쇄 영역.
                var probe = doc.createElement('div');
                probe.style.cssText = 'position:absolute;visibility:hidden;width:182mm;height:269mm;';
                doc.body.appendChild(probe);
                var box = probe.getBoundingClientRect();
                var pageW = box.width, pageH = box.height;
                doc.body.removeChild(probe);

                page.style.width = pageW + 'px';
                page.style.zoom = 1;
                // getBoundingClientRect는 zoom이 반영된 "실제 그려지는" 높이라
                // scrollHeight(브라우저마다 zoom 반영 여부가 다름)보다 안전하다.
                var naturalH = page.getBoundingClientRect().height;
                if (naturalH > pageH) {{
                    page.style.zoom = (pageH / naturalH) * 0.99;  // 0.99: 반올림 여유
                }}
            }}
            w.focus();
            w.print();
        }};
        // 폰트가 다 로드되기 전에 재면 글자 크기가 달라져 높이 계산이 통째로 틀어진다.
        if (w.document.fonts && w.document.fonts.ready) {{
            w.document.fonts.ready.then(function () {{ setTimeout(fitAndPrint, 60); }});
        }} else {{
            setTimeout(fitAndPrint, 350);
        }}
    }};
    </script>
    """
    components.html(button_html, height=54)


# ------------------------------------------------------------------ 기본 설정
st.set_page_config(page_title="현장안전 통합분석 대시보드", page_icon="🦺", layout="wide")

# ------------------------------------------------------------------ 세방(주) 브랜드 아이덴티티
# 배경/글자/사이드바/버튼·탭 강조색과 세방고딕 폰트는 .streamlit/config.toml의 네이티브
# [theme]/[theme.sidebar]/[[theme.fontFaces]] 설정으로 적용한다(다크모드 자동감지와
# 충돌하는 수동 CSS 주입 대신 — 그러면 Streamlit이 라이트/다크 전반에 일관되게 반영해준다).
# 주의: 폭염 단계 신호색(heatwave.py의 LEVEL_COLOR/NORMAL_COLOR, 정상=초록~위험=빨강)은
# 안전 신호등 의미라 브랜드색 적용 대상이 아니다 — 이 테마와 무관하게 그대로 유지.
SEBANG_DARK_GRAY = "#3B4551"     # Primary: Pantone 432 C
SEBANG_LIGHT_GRAY_1 = "#E4E4E2"  # Primary: Pantone Cool Gray 2 C
SEBANG_LIGHT_GRAY_2 = "#A7A9AC"  # Primary: Pantone 429 C
SEBANG_GREEN = "#009CA6"         # Secondary: Pantone 320 C
SEBANG_ORANGE = "#EE2737"        # Secondary: Pantone 1788 C

# 점검 결과가 "양호"(정상 응답)인 칸에 칠하는 옅은 초록 — 이슈 강조색(#c81d25)과 달리
# 글자를 검게 두어, 표를 훑을 때 빨강만 눈에 띄고 초록은 배경으로 물러나게 한다.
_OK_CELL_STYLE = "background-color:#d9f2e3; color:#14532d"

_FONT_DIR = Path(__file__).parent / "fonts"
_SEBANG_REGULAR = _FONT_DIR / "SEBANG Gothic.ttf"
_SEBANG_BOLD = _FONT_DIR / "SEBANG Gothic Bold.ttf"


def _check_password() -> bool:
    """APP_PASSWORD(secrets)가 설정돼 있으면 맞는 비밀번호를 입력해야 통과.

    secrets 자체가 없는(로컬에서 아직 설정 안 한) 환경에서는 개발 편의상 그냥 통과시킨다 —
    배포판(Streamlit Cloud)엔 반드시 secrets로 APP_PASSWORD를 등록해야 실제로 보호된다.
    """
    try:
        required = st.secrets.get("APP_PASSWORD", "")
    except Exception:
        required = ""
    if not required:
        return True
    if st.session_state.get("_authed"):
        return True

    st.title("🦺 현장안전 통합분석 대시보드")
    pw = st.text_input("접속 비밀번호", type="password")
    if pw:
        if pw == required:
            st.session_state["_authed"] = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")
    return False


if not _check_password():
    st.stop()


def _analysis_unlocked() -> bool:
    """분석 탭(전사 현황/지사 상세/점검 편향분석) 전용 2차 비밀번호 확인.

    폭염 대응 탭은 지사 전체가 매주 쓰는 운영 화면이라 APP_PASSWORD만으로 열리지만,
    분석 탭 3개는 연구용 원자료(점검·재해 이력 전수)라 별도 비밀번호를 하나 더 둔다.
    ANALYSIS_PASSWORD가 secrets에 없으면 잠그지 않는다(로컬 개발 편의).
    """
    try:
        required = st.secrets.get("ANALYSIS_PASSWORD", "")
    except Exception:
        required = ""
    return (not required) or bool(st.session_state.get("_analysis_authed"))


def _analysis_gate(tab_key: str) -> bool:
    """잠긴 분석 탭 안에 비밀번호 입력창을 그리고, 통과했으면 True.

    탭마다 위젯 key가 달라야 해서 tab_key를 받는다(같은 key를 쓰면 Streamlit이
    중복 위젯으로 보고 에러를 낸다).
    """
    if _analysis_unlocked():
        return True
    st.info("🔒 이 탭은 연구용 원자료를 포함하고 있어 별도 비밀번호가 필요합니다.")
    pw = st.text_input("분석 탭 비밀번호", type="password", key=f"_analysis_pw_{tab_key}")
    if pw:
        if pw == st.secrets.get("ANALYSIS_PASSWORD", ""):
            st.session_state["_analysis_authed"] = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")
    return False

# 차트(matplotlib)도 세방고딕을 우선 쓰고, 파일이 없는 예외 상황에서만 번들 폰트/윈도우 폰트로 폴백.
_FONT_CANDIDATES = [
    _SEBANG_REGULAR,
    Path(__file__).parent / "fonts" / "NanumGothic.ttf",
    Path(r"C:\Windows\Fonts\malgun.ttf"),
]
FONT_PATH = next((str(p) for p in _FONT_CANDIDATES if p.exists()), None)
if FONT_PATH:
    try:
        fm.fontManager.addfont(FONT_PATH)
        plt.rcParams["font.family"] = fm.FontProperties(fname=FONT_PATH).get_name()
        plt.rcParams["axes.unicode_minus"] = False
    except Exception:
        FONT_PATH = None

# 차트 강조색은 브랜드 그린으로 통일(급증 워드클라우드·게이지 스텝 등 의미색은 그대로 둠).
CHART_ACCENT = SEBANG_GREEN


@st.cache_data
def load():
    return S.load_data()

insp, acc = load()
INSP_BRANCHES = sorted(insp["지사"].unique())

# ------------------------------------------------------------------ 사이드바
st.sidebar.title("🦺 현장안전 통합분석")
st.sidebar.caption("세방(주) · 프로토타입")

# 브라우저 F5(하드 새로고침)는 새 세션을 만들어 _check_password()의 session_state가
# 초기화되므로 비밀번호를 다시 묻는다. 이 버튼은 스크립트만 재실행(st.rerun)해서
# 같은 세션을 유지한 채 캐시된 데이터만 최신화한다.
if st.sidebar.button("🔄 새로고침", width="stretch",
                      help="비밀번호 재입력 없이 화면과 데이터만 최신 상태로 다시 불러옵니다."):
    st.cache_data.clear()
    st.rerun()

period_opt = st.sidebar.radio("분석 기간", ["최근 90일", "최근 180일", "전체 기간"], index=2)
end = insp["점검일자"].max()
if period_opt == "최근 90일":
    period = (end - pd.Timedelta(days=90), end)
elif period_opt == "최근 180일":
    period = (end - pd.Timedelta(days=180), end)
else:
    period = None

SEV_COL_MAP = {"규칙기반(라벨)": "심각도점수", "LLM 하이브리드(맥락반영·파일럿)": "LLM심각도"}
sev_label = st.sidebar.radio(
    "심각도 산정 기준", list(SEV_COL_MAP.keys()),
    help="LLM 하이브리드는 195건 직접 의미판정 + 590건 키워드 맥락보정을 결합한 파일럿 결과입니다.")
SEV_COL = SEV_COL_MAP[sev_label]

st.sidebar.markdown("---")
st.sidebar.caption(
    f"데이터: 점검 {len(insp):,}건 ({insp['점검일자'].min().date()}~{insp['점검일자'].max().date()}) · "
    f"인적재해 이력 {int((acc['재해성격']=='인적재해').sum()):,}건")

SYNC_SUMMARY_PATH = Path(__file__).parent / "data" / "_last_sync_summary.json"

# gsafety.kr 로그인 자격증명이 있는 로컬 환경에서만 동기화 버튼을 보여준다.
# 클라우드 배포판은 이 자격증명도, Downloads 폴더의 원본 파일도, 사내망 접근도 없어
# 눌러도 100% 실패하므로 아예 노출하지 않는다.
if os.environ.get("GSAFETY_ID") and os.environ.get("GSAFETY_PW"):
    if st.sidebar.button("🔄 최신 점검 데이터 동기화", width="stretch",
                          help="gsafety.kr에서 지사 점검 데이터를 가져와 누적본에 반영합니다(수 분 소요)"):
        with st.sidebar:
            with st.spinner("사이트 로그인 → 신규 데이터 조회 → 정제·마스킹 재실행 중..."):
                result = subprocess.run(
                    [sys.executable, str(Path(__file__).parent / "sync_gsafety.py")],
                    capture_output=True, text=True, encoding="utf-8", errors="replace",
                    env={**os.environ, "PYTHONIOENCODING": "utf-8"},
                )
        if result.returncode == 0:
            summary = {}
            if SYNC_SUMMARY_PATH.exists():
                summary = json.loads(SYNC_SUMMARY_PATH.read_text(encoding="utf-8"))
            st.session_state["sync_result"] = ("success", summary)
            st.cache_data.clear()
        else:
            st.session_state["sync_result"] = ("error", (result.stdout + result.stderr)[-2000:])
        st.rerun()

if "sync_result" in st.session_state:
    kind, payload = st.session_state.pop("sync_result")
    with st.sidebar:
        if kind == "error":
            st.error("동기화 실패")
            st.code(payload, language=None)
        else:
            by_branch = payload.get("by_branch", {})
            if not by_branch:
                st.info("동기화 완료 — 신규 데이터 없음(이미 최신 상태)")
            else:
                st.success(f"동기화 완료 — 신규 {payload.get('total_added', 0)}건 반영")
                rows = []
                for branch, cnt in sorted(by_branch.items(), key=lambda x: -x[1]):
                    b_acc = acc[acc["지사"] == branch]
                    rows.append({
                        "지사": branch,
                        "신규 점검활동": cnt,
                        "인적재해(누적)": int((b_acc["재해성격"] == "인적재해").sum()),
                        "재물·차량사고(누적)": int((b_acc["재해성격"] == "재물·차량사고").sum()),
                    })
                st.dataframe(pd.DataFrame(rows), hide_index=True, width="stretch")

# ------------------------------------------------------------------ 공통 계산
ri = S.branch_risk_index(insp, period, severity_col=SEV_COL)

# ================================================================== 헤더
st.title("현장 안전 통합 상황판")
_mode = "🧪 LLM 하이브리드(파일럿)" if SEV_COL == "LLM심각도" else "📐 규칙기반(라벨)"
st.caption(f"심각도 기준: **{_mode}** · 분석기간: {period_opt}")

tab4, tab_briefing, tab1, tab2, tab3 = st.tabs(
    ["🌡️ 폭염 대응", "📰 일일 안전보건 브리핑", "📊 전사 현황", "🏢 지사 상세", "🔍 점검 편향분석"])

# ================================================================== TAB4 폭염 대응
with tab4:
    if not HW.available():
        st.warning("아직 관측 데이터가 없습니다. `온도를 조져보자` 프로젝트의 main.py를 먼저 실행해주세요.")
    else:
        obs = HW.load_observations()
        notif = HW.load_notifications()

        if obs.empty:
            st.info("아직 쌓인 관측 데이터가 없습니다. 스케줄러가 최소 1회 이상 실행된 뒤 다시 확인해주세요.")
        else:
            # 스케줄러(사내 PC)가 멈추면 관측 DB가 갱신되지 않는데, 화면은 옛 온도를
            # 현재값처럼 계속 보여준다 — 폭염 판단을 잘못하게 되므로 눈에 띄게 알린다.
            _latest_obs = obs["observed_at"].max()
            if pd.notna(_latest_obs):
                _age_h = (HW.now_kst() - _latest_obs).total_seconds() / 3600
                if _age_h >= 3:
                    st.error(
                        f"⚠️ 관측 데이터가 **{_age_h:.0f}시간째 갱신되지 않았습니다** "
                        f"(최신 관측 {_latest_obs.strftime('%m-%d %H:%M')}). 화면의 체감온도는 "
                        "현재 상태가 아닐 수 있습니다 — 자동 수집 스케줄러가 멈췄는지 확인해주세요.",
                        icon="⚠️")

            _sheet_code, _sheet_msg = HW.incident_sheet_status()
            if _sheet_code != "ok":
                st.error(f"⚠️ 지사 제출 데이터를 읽지 못하고 있습니다 — {_sheet_msg} "
                         "아래 표의 0은 '제출 없음'이 아니라 **연동 오류**일 수 있습니다.", icon="⚠️")

            # CSO/대표가 지난주(또는 그 이전) 보고서를 다시 요청할 수 있어, 인쇄 시점의
            # "이번주"가 아니라 원하는 주차를 직접 골라 그 주 데이터로 다시 뽑을 수 있게
            # 한다 — 선택 안 하면 기본값(이번주)으로 지금까지와 동일하게 동작한다.
            _report_weeks = HW.available_report_weeks()
            _report_week_labels = [HW.week_label(w) for w in _report_weeks]
            _report_week_idx = st.selectbox(
                "🗓️ 인쇄할 보고서 주차", range(len(_report_weeks)),
                format_func=lambda i: _report_week_labels[i], index=0,
                help="기본값은 이번주입니다. 지난 주차를 고르면 그 주 데이터로 보고서를 다시 뽑습니다.")
            _render_print_report_button(_report_weeks[_report_week_idx])
            st.markdown("---")

            clusters = HW.map_clusters(obs)

            col_map, col_kpi = st.columns([3, 2])

            with col_map:
                st.markdown('<div id="mapColAnchor"></div>', unsafe_allow_html=True)
                st.subheader("🗺️ 지사별 폭염 현황 지도")
                kakao_key = st.secrets.get("KAKAO_JS_KEY", "")
                if not kakao_key:
                    st.info("카카오맵 API 키(KAKAO_JS_KEY)가 secrets에 없어 지도를 표시할 수 없습니다. "
                            "`.streamlit/secrets.toml`(로컬) 또는 Streamlit Cloud의 Secrets 설정에 등록해주세요.")
                elif not clusters:
                    st.info("사업장 좌표가 없습니다. `config/sites.yaml`을 확인해주세요.")
                else:
                    def _marker_detail(d: dict) -> str:
                        if not d["has_data"]:
                            return "관측 데이터 없음"
                        prefix = "~" if d.get("is_estimate") else ""
                        if d.get("is_estimate"):
                            # estimate_km은 DB에 저장된 추정치(목포/당진/삼천포 등)에는 아직
                            # 없을 수 있어(estimate_source 문자열에 거리가 이미 포함돼 있음)
                            # None이면 생략한다 — 여기서 죽으면 폭염 탭 전체가 크래시난다.
                            km = d.get("estimate_km")
                            km_part = f' {km:.0f}km' if km is not None else ""
                            text = (f'{prefix}{d["apparent_temp"]:.1f}℃ 추정 (자체 관측지점 없음, '
                                    f'{d["estimate_source"]}{km_part} 값 사용)')
                        else:
                            text = f'{prefix}{d["apparent_temp"]:.1f}℃ ({HW.level_label(d["level"])})'
                        if d.get("official_advisory"):
                            text += f' · 기상청 공식 폭염{d["official_advisory"]} 발효중'
                        return text

                    def _marker_image(color: str, badge: str, label: str, is_estimate: bool, is_office: bool,
                                       is_advisory: bool = False):
                        """원(체감온도 배지) + 아래 라벨 칩을 SVG 하나로 그려 data URI로 반환.

                        카카오맵 CustomOverlay(임의 HTML 오버레이)가 이 배포 환경에서 DOM에 붙지
                        않는 문제가 있어(마커 자체는 정상 동작 확인됨), 오버레이 대신 SVG를 구운
                        마커 이미지로 대체했다 — 렌더링을 브라우저의 기본 <img> 디코딩에 맡기므로
                        더 안정적이다. 반환값: (data URI, 전체폭, 전체높이, 앵커x, 앵커y) — 앵커는
                        원의 중심(=실제 좌표가 가리키는 지점)이다.

                        is_advisory=True면 폭염특보 사이렌(🚨)을 라벨 텍스트 앞이 아니라 원 오른쪽
                        위쪽에 작은 배지로 그린다 — 지도에 마커가 촘촘히 찍혀 있을 때 라벨 텍스트
                        속에 묻히지 않고 한눈에 띄도록.
                        """
                        d = 42 if is_office else 28
                        font_size = 14 if is_office else 10
                        label_font = 10
                        label_w = max(d + 6, len(label) * 9 + 10)
                        label_h = 15
                        gap = 3
                        w, h = max(d, label_w), d + gap + label_h
                        cx, cy = w / 2, d / 2
                        r = d / 2 - 1.5
                        # 사이렌 배지가 원 바깥(오른쪽 위)으로 튀어나오는 만큼 캔버스에 여백을
                        # 더해준다 — 안 그러면 잘려 보인다. 기존 원·라벨 좌표는 아래로 pad만큼
                        # 그대로 밀어서(y_shift) 앵커(=실제 좌표 지점)만 새 위치로 갱신하면 된다.
                        badge_r = 7
                        pad = badge_r + 3 if is_advisory else 0
                        w = w + pad  # 가로는 원 중심(cx) 유지, 오른쪽 여백만 늘어남
                        h = h + pad
                        cy = cy + pad
                        dash = ' stroke-dasharray="4,3"' if is_estimate else ""
                        badge_e, label_e = html.escape(badge), html.escape(label)
                        advisory_svg = ""
                        if is_advisory:
                            bcx, bcy = cx + r * 0.72, cy - r * 0.72
                            advisory_svg = (
                                f'<circle cx="{bcx}" cy="{bcy}" r="{badge_r}" fill="white" '
                                f'stroke="#c81d25" stroke-width="1.5"/>'
                                f'<text x="{bcx}" y="{bcy + 1}" text-anchor="middle" dominant-baseline="central" '
                                f'font-size="10">🚨</text>'
                            )
                        svg = (
                            f'<svg xmlns="http://www.w3.org/2000/svg" width="{w}" height="{h}">'
                            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="{color}" '
                            f'stroke="white" stroke-width="2"{dash}/>'
                            f'<text x="{cx}" y="{cy}" text-anchor="middle" dominant-baseline="central" '
                            f'font-size="{font_size}" font-weight="700" fill="white" '
                            f'font-family="sans-serif">{badge_e}</text>'
                            f'{advisory_svg}'
                            f'<rect x="{cx - label_w / 2}" y="{d + gap + pad}" width="{label_w}" height="{label_h}" '
                            f'rx="3" fill="white" fill-opacity="0.85"/>'
                            f'<text x="{cx}" y="{d + gap + label_h / 2 + pad}" text-anchor="middle" '
                            f'dominant-baseline="central" font-size="{label_font}" font-weight="700" '
                            f'fill="#111" font-family="sans-serif">{label_e}</text>'
                            f'</svg>'
                        )
                        uri = "data:image/svg+xml;base64," + base64.b64encode(svg.encode("utf-8")).decode("ascii")
                        return uri, round(w), round(h), round(cx), round(cy)

                    # 실제 좌표에 마커를 찍는 진짜 지도라, 예전 정적 SVG 버전에서 필요했던
                    # "겹치지 않게 픽셀 단위로 손으로 미는" 보정이 더 이상 필요 없다 — 카카오맵
                    # 자체 줌/팬으로 사용자가 겹친 구간을 직접 확대해서 볼 수 있다.
                    markers = []
                    for c in clusters:
                        badge = f'{"~" if c.get("is_estimate") else ""}{c["apparent_temp"]:.0f}°' if c["has_data"] else "–"
                        label = c["branch"]
                        img, w, h, ax, ay = _marker_image(HW.level_color(c["level"]), badge, label,
                                                           bool(c.get("is_estimate")), True,
                                                           is_advisory=bool(c.get("official_advisory")))
                        markers.append({
                            "lat": c["lat"], "lon": c["lon"], "img": img, "w": w, "h": h, "ax": ax, "ay": ay,
                            "title": f'{c["branch"]} {c["city"]}(사무실) · {_marker_detail(c)}',
                        })
                        for s in c["satellites"]:
                            # 창원(진해, 부산 부속)는 경남 지사 사무실과 좌표가 같아져서(2026-08-02
                            # 좌표 정정) 라벨이 겹쳐 보인다 — 지도 표시만 생략(체감온도 계산·부산
                            # 지사 대표값 산정 등 다른 로직에는 계속 사용됨).
                            if s["city"] == "창원(진해)":
                                continue
                            s_badge = f'{"~" if s.get("is_estimate") else ""}{s["apparent_temp"]:.0f}°' if s["has_data"] else "–"
                            s_img, s_w, s_h, s_ax, s_ay = _marker_image(
                                HW.level_color(s["level"]), s_badge, s["city"], bool(s.get("is_estimate")), False)
                            markers.append({
                                "lat": s["lat"], "lon": s["lon"], "img": s_img, "w": s_w, "h": s_h,
                                "ax": s_ax, "ay": s_ay,
                                "title": f'{c["branch"]} {s["city"]}(부속, {s["site_count"]}개소) · {_marker_detail(s)}',
                            })

                    # 마커 JSON을 <script> 안에 그대로 박아 넣으므로, 값에 우연히 "</script>"가
                    # 섞여 있어도 태그가 조기 종료되지 않도록 이스케이프한다.
                    markers_json = json.dumps(markers, ensure_ascii=False).replace("</", "<\\/")

                    map_html = f"""
                    <meta http-equiv="Content-Security-Policy" content="upgrade-insecure-requests">
                    <div id="kakaoMap" style="width:100%; height:1050px; border-radius:8px;"></div>
                    <script src="https://dapi.kakao.com/v2/maps/sdk.js?appkey={kakao_key}&autoload=false"></script>
                    <script>
                    kakao.maps.load(function () {{
                        var container = document.getElementById('kakaoMap');
                        var map = new kakao.maps.Map(container, {{
                            center: new kakao.maps.LatLng(36.2, 127.9), level: 12,
                        }});
                        map.addControl(new kakao.maps.ZoomControl(), kakao.maps.ControlPosition.RIGHT);

                        var markers = {markers_json};
                        var bounds = new kakao.maps.LatLngBounds();
                        markers.forEach(function (m) {{
                            var pos = new kakao.maps.LatLng(m.lat, m.lon);
                            bounds.extend(pos);
                            var image = new kakao.maps.MarkerImage(
                                m.img, new kakao.maps.Size(m.w, m.h),
                                {{offset: new kakao.maps.Point(m.ax, m.ay)}}
                            );
                            new kakao.maps.Marker({{position: pos, map: map, image: image, title: m.title}});
                        }});
                        // 마커 bounds에 딱 맞추면 국토가 세로로 긴 형태라(경도 폭 < 위도 폭)
                        // 가로로 넓은 컨테이너에서는 좌우로 바다만 넓게 남는다. 지도 높이를
                        // 늘려(kakaoMap div height=1050px) 세로 비율을 국토 형태에 가깝게
                        // 맞춰 여백 없이 국토 위주로 채운다.
                        function fitMap() {{
                            map.relayout();
                            map.setBounds(bounds);
                        }}
                        fitMap();
                        // Streamlit의 components.html iframe은 로드 직후엔 부모가 아직
                        // 최종 크기로 자리잡기 전이라, 카카오맵이 그 순간의 좁은 크기로
                        // 초기화된 채 멈춰(실제 칸은 넓어졌는데 지도만 왼쪽 위 구석에 작게
                        // 그려지고 나머지는 빈 공간으로 남는 증상) 버리는 경우가 있다.
                        // 컨테이너 크기가 바뀔 때마다 relayout+bounds를 다시 계산해 항상
                        // 실제 렌더링 크기에 맞춘다.
                        new ResizeObserver(fitMap).observe(container);
                    }});
                    </script>
                    """
                    components.html(map_html, height=1065, scrolling=False)
                    legend_items = [(None, "정상"), ("주의", "주의 33℃+"),
                                     ("경고", "경고 35℃+"), ("위험", "위험 38℃+")]
                    legend_html = "".join(
                        f'<span style="display:inline-flex; align-items:center; margin-right:18px;">'
                        f'<span style="display:inline-block; width:10px; height:10px; border-radius:50%; '
                        f'background:{HW.level_color(lvl)}; margin-right:6px;"></span>{label}</span>'
                        for lvl, label in legend_items
                    )
                    st.markdown(
                        f'<div style="display:flex; flex-wrap:wrap; align-items:center; margin-top:6px;">'
                        f'{legend_html}'
                        f'<span style="font-size:12px; color:#666;">🚨 = 폭염특보 발생</span></div>',
                        unsafe_allow_html=True)

                    latest_obs = obs["observed_at"].max() if not obs.empty else None
                    st.caption(
                        "출처 : 기상청 API 허브 자료"
                        + (f" · 최신 관측 시간 : {latest_obs.strftime('%m-%d %H:%M')} 기준"
                           if pd.notna(latest_obs) else " · 관측 데이터 없음")
                        + " · 체감온도는 기상청 여름철 체감온도 공식으로 산출한 추정치이며, "
                          "사업장의 실측 온도와 상이할 수 있습니다."
                    )

            with col_kpi:
                # 예전엔 고정 높이 컨테이너(height=615)를 썼다가 내용이 넘칠 때 내부
                # 스크롤이 생겨 어색해서, 높이를 내용에 맡기는 쪽으로 바꿨었다. 그런데
                # 그러면 반대로 지도(고정 900~1050px)보다 이 카드가 짧을 땐 그 아래가
                # 빈 여백으로 남고, 내용이 많을 땐 지도보다 카드가 길어져 옆으로 나란히
                # 볼 때 비율이 안 맞아 보인다는 피드백을 받았다 — 아래 fit 스크립트로
                # 카드 높이를 지도와 항상 맞추고, 내용이 넘치면 스크롤 대신 CSS zoom으로
                # 전체를 살짝 축소해 한 화면 안에 다 들어오게 한다.
                with st.container(border=True):
                    st.subheader("🔥 현재 최고 체감온도 지사")
                    summary_now = HW.branch_summary(obs)
                    hot = summary_now[summary_now["has_data"]] if not summary_now.empty else summary_now
                    if hot.empty:
                        st.info("관측 데이터가 있는 지사가 아직 없습니다.")
                    else:
                        top = hot.loc[hot["apparent_temp"].idxmax()]
                        est_prefix = "~" if top["is_estimate"] else ""
                        st.metric(
                            label=f'{top["branch"]} · {top["worst_city"]}',
                            value=f'{est_prefix}{top["apparent_temp"]:.1f}℃',
                        )
                        st.caption(f'{HW.level_label(top["level"])} 단계'
                                   + (" · 자체 관측지점 없어 추정치" if top["is_estimate"] else "")
                                   + (f' · 🚨 기상청 공식 폭염{top["official_advisory"]} 발효중'
                                      if top.get("official_advisory") else ""))

                    st.divider()
                    st.markdown("##### 🏥 사업장 온열질환 환자 발생 현황")
                    # 공식 집계는 "발생 즉시 보고" 폼(사건 단위). 주차별 폼의 자기보고는
                    # 교차검증용으로만 쓴다 — 두 값이 다르면 아래에서 지사를 짚어준다.
                    esum = HW.patient_event_summary()
                    psum = HW.patient_summary()
                    today_label = HW.now_kst().strftime("%y.%m.%d")
                    th_style = (f"background:{SEBANG_DARK_GRAY}; color:white; padding:8px 4px; font-size:13px; "
                                f"border:1px solid {SEBANG_DARK_GRAY};")
                    td_style = (f"background:{SEBANG_LIGHT_GRAY_1}; color:{SEBANG_DARK_GRAY}; padding:12px 4px; "
                                f"font-size:16px; font-weight:700; border:1px solid {SEBANG_LIGHT_GRAY_2};")
                    st.markdown(
                        '<table style="width:100%; border-collapse:collapse; text-align:center; '
                        'font-family:\'SEBANG Gothic\',sans-serif; margin-bottom:8px;">'
                        f'<tr><th style="{th_style}">25년(전년도)</th>'
                        f'<th style="{th_style}">26년(누적)</th>'
                        f'<th style="{th_style}">이번주</th></tr>'
                        f'<tr><td style="{td_style}">0명</td>'
                        f'<td style="{td_style}">{esum["cumulative"]}명</td>'
                        f'<td style="{td_style}">{esum["this_week"]}명</td></tr>'
                        '</table>',
                        unsafe_allow_html=True,
                    )
                    _events = HW.weekly_patient_events()
                    if not _events.empty:
                        for _e in _events.itertuples():
                            st.error(f"**{_e.branch}** {int(_e.환자수)}명 · {_e.증상 or '증상 미기재'} "
                                     f"· {_e.timestamp.strftime('%m-%d %H:%M')}")
                            if _e.상세:
                                st.text(_e.상세)
                    _mis = HW.patient_report_mismatch()
                    if not _mis.empty:
                        _rows = ", ".join(
                            f"{r.branch}(즉시 {r.즉시보고}명 / 주간 {r.주간보고}명)" for r in _mis.itertuples())
                        st.warning(f"⚠️ 즉시보고와 주간보고 환자 수가 다른 지사: {_rows} — 어느 쪽을 "
                                   "누락했는지 해당 지사에 확인이 필요합니다.", icon="⚠️")
                    st.link_button("🏥 온열질환 발생 보고하기", _PATIENT_FORM_URL, width="stretch")

                    st.divider()
                    st.markdown("##### 🚨 작업조정·중지 보고")
                    ssum = HW.stoppage_summary()
                    st.markdown(
                        '<table style="width:100%; border-collapse:collapse; text-align:center; '
                        'font-family:\'SEBANG Gothic\',sans-serif; margin-bottom:8px;">'
                        f'<tr><th style="{th_style}">\'26년</th>'
                        f'<th style="{th_style}">이번주</th>'
                        f'<th style="{th_style}">금일({today_label})</th></tr>'
                        f'<tr><td style="{td_style}">{ssum["season_cumulative"]}건</td>'
                        f'<td style="{td_style}">{ssum["this_week"]}건</td>'
                        f'<td style="{td_style}">{ssum["today"]}건</td></tr>'
                        '</table>',
                        unsafe_allow_html=True,
                    )
                    stoppages = HW.today_stoppages()
                    if stoppages.empty:
                        st.info("금일 작업중지·조정 보고 없음")
                    else:
                        # "작업 조정·중지 즉시 보고"는 사건 1건당 1행이라, 오늘 같은 지사가
                        # 여러 건 제출했으면 대표값 하나로 접지 않고 건마다 카드를 그대로 보여준다.
                        for row in stoppages.itertuples():
                            # 상세는 구글폼 자유서술/선택형 입력이라, 마크다운 문법(링크 등)이
                            # 그대로 해석되지 않도록 본문(st.error)과 분리해 일반 텍스트로만 표시한다.
                            detail = row.상세 or "(상세 미기재)"
                            icon = "🚨" if "중지" in row.구분 else "🟡"
                            st.error(f"**{row.branch}** {icon} {row.구분} · {row.timestamp.strftime('%H:%M')}")
                            st.text(detail)
                    st.link_button("🚨 지금 작업조정·중지 보고하기", _STOPPAGE_FORM_URL, width="stretch")

                # 카드(위 st.container(border=True))를 지도 컬럼과 같은 높이로 강제하고,
                # 내용이 넘치면 스크롤이 아니라 CSS zoom으로 전체를 살짝 줄여 한 화면에
                # 다 들어오게 한다. 카드 "안"이 아니라 col_kpi 레벨에 둬서, 이 스크립트
                # 자신(높이 0 iframe)이 카드의 내용 높이 측정에 섞여 들어가지 않게 한다.
                components.html("""
                <script>
                (function () {
                    function fit() {
                        var doc = window.parent.document;
                        var anchor = doc.getElementById('mapColAnchor');
                        if (!anchor) return;
                        var mapCol = anchor.closest('[data-testid="stColumn"]');
                        if (!mapCol) return;
                        var row = mapCol.parentElement;
                        var kpiCol = row && row.children[1];
                        if (!kpiCol) return;
                        var card = null;
                        kpiCol.querySelectorAll('[data-testid="stVerticalBlock"]').forEach(function (el) {
                            if (!card && parseFloat(getComputedStyle(el).borderWidth) > 0) card = el;
                        });
                        if (!card) return;
                        var targetH = mapCol.scrollHeight;
                        // Streamlit 자체 CSS(emotion)가 stVerticalBlock의 높이를 !important로
                        // 깔아둬서, 그냥 .style.height = ... 로는 무시된다(직접 확인함) —
                        // setProperty에 'important'를 명시해야 실제로 적용된다.
                        card.style.setProperty("zoom", "1", "important");
                        card.style.setProperty("height", targetH + "px", "important");
                        card.style.setProperty("min-height", targetH + "px", "important");
                        card.style.setProperty("max-height", targetH + "px", "important");
                        card.style.setProperty("overflow", "hidden", "important");
                        var naturalH = card.scrollHeight;
                        if (naturalH > targetH) {
                            card.style.setProperty("zoom", Math.max(0.5, targetH / naturalH), "important");
                        }
                    }
                    fit();
                    setTimeout(fit, 400);
                    setTimeout(fit, 1200);
                    try {
                        new ResizeObserver(fit).observe(window.parent.document.body);
                    } catch (e) {}
                })();
                </script>
                """, height=0)

            st.markdown("---")
            st.subheader("📅 주차별 온열질환 대응 현황")
            if not HW.GOOGLE_SHEET_CSV_URL:
                st.warning("⚠️ 구글폼 응답 시트가 아직 연결되지 않아 전부 0으로 표시됩니다. "
                           "`heatwave.py`의 GOOGLE_SHEET_CSV_URL을 채우면 실제 제출값으로 바뀝니다.", icon="⚠️")
            weekly_incident = HW.weekly_incident_totals()
            if not weekly_incident.empty:
                weekly_incident["이슈"] = ((weekly_incident["환자수"] > 0) | (weekly_incident["작업조정"] > 0)
                                          | (weekly_incident["작업중지"] > 0))
                weekly_incident = weekly_incident.sort_values(
                    ["이슈", "환자수", "작업중지", "작업조정"], ascending=[False, False, False, False])
                display_df = weekly_incident[
                    ["branch", "환자수", "작업조정", "작업중지", "최근제출"]
                ].rename(columns={"branch": "지사"})
                display_df["환자수"] = display_df["환자수"].map(lambda n: f"{int(n)}명")
                display_df["작업조정"] = display_df["작업조정"].map(lambda n: f"{int(n)}건")
                display_df["작업중지"] = display_df["작업중지"].map(lambda n: f"{int(n)}건")
                display_df["최근제출"] = display_df["최근제출"].apply(
                    lambda t: t.strftime("%m-%d %H:%M") if pd.notna(t) else "이번주 제출 없음")
                display_df["보고"] = display_df["지사"].map(_branch_form_url)

                # 예전엔 이슈(환자/조정/중지 발생)가 있으면 행 전체를 진한 빨강으로
                # 칠했는데, 글자가 안 보일 정도로 가독성이 나빠 셀 단위로 옅게 바꿨다
                # — 이슈가 있는 칸만 옅은 빨강으로 표시한다.
                _today_wd = HW.now_kst().weekday()  # 0=월 ... 3=목, 4=금
                _INCIDENT_CELL_STYLE = "background-color:#fdd7d7; color:#900; font-weight:700"

                def _highlight_issue(row):
                    cols = list(row.index)
                    styles = [""] * len(row)
                    not_submitted = row["최근제출"] == "이번주 제출 없음"
                    for field in ("환자수", "작업조정", "작업중지"):
                        zero_label = "0명" if field == "환자수" else "0건"
                        if row[field] != zero_label:
                            styles[cols.index(field)] = _INCIDENT_CELL_STYLE
                    # 이번주 아직 제출이 없으면 요일이 늦어질수록(목요일 주황 → 금요일
                    # 빨강) "최근 제출" 칸만 경고색으로 강조해 미제출 독촉이 눈에 띄게 한다.
                    if not_submitted:
                        if _today_wd == 4:
                            styles[cols.index("최근제출")] = "background-color:#c81d25; color:white; font-weight:700"
                        elif _today_wd == 3:
                            styles[cols.index("최근제출")] = "background-color:#f2a154; color:white; font-weight:700"
                    else:
                        # 전부 0이고 이번주 제출도 있으면 "확인된 양호" — 행 전체를 초록으로.
                        if row["환자수"] == "0명" and row["작업조정"] == "0건" and row["작업중지"] == "0건":
                            styles = [_OK_CELL_STYLE] * len(row)
                    return styles

                st.dataframe(
                    display_df.style.apply(_highlight_issue, axis=1),
                    width="stretch", hide_index=True,
                    column_config={
                        "최근제출": st.column_config.TextColumn("최근 제출"),
                        "보고": st.column_config.LinkColumn("보고", display_text="📝 보고하기", width="small"),
                    },
                )
                st.caption("환자수/작업조정 = 매일 신규 건수 문항을 **이번주 날짜별로 합산**(같은 날 "
                           "여러 번 제출해도 그날 마지막 값만 대표값으로 써서 중복 합산 방지) — 누적 "
                           "계산은 응답자가 아니라 대시보드가 담당합니다. 작업중지 = 이번주 \"작업중지 "
                           "즉시 보고\" 제출 건수(사건 1건당 1건). 옅은 빨강 칸 = 그 항목이 0보다 큼"
                           "(맨 위로 정렬). 초록 행 = 이번주 제출이 있고 전부 0인 지사(양호) — 아직 "
                           "제출하지 않은 지사는 0이어도 \"확인된 양호\"가 아니라 색을 칠하지 않습니다. "
                           "\"최근 제출\" 칸은 이번주 미제출 지사에 한해 목요일부터 주황, 금요일부터 빨강으로 "
                           "독촉 표시됩니다. \"최근 제출\" = 두 폼 중 이번주 그 지사의 마지막 제출 시각. "
                           "\"보고\" = 클릭하면 그 지사가 선택된 채로 온열질환·예방점검 구글폼이 새 탭에서 "
                           "열립니다. 작업중지 상세 내용은 위 \"🚨 작업조정·중지 보고\" 카드나 인쇄 보고서를 "
                           "확인하세요.")

            st.markdown("---")
            st.subheader("🧾 주간 온열질환 예방 점검 결과")
            checklist = HW.weekly_checklist()
            if not checklist.empty:
                label_cols = list(HW.CHECKLIST_LABELS.values())

                def _is_issue(v) -> bool:
                    return "특이사항 기재" in str(v)

                checklist["이슈"] = checklist[HW.CHECKLIST_FIELDS].apply(
                    lambda row: any(_is_issue(v) for v in row), axis=1)
                checklist = checklist.sort_values("이슈", ascending=False)
                display_chk = checklist[
                    ["branch", *HW.CHECKLIST_FIELDS, "점검특이사항", "최근제출", "제출횟수"]
                ].rename(columns={"branch": "지사", "점검특이사항": "특이사항", **HW.CHECKLIST_LABELS})
                display_chk["최근제출"] = checklist["최근제출"].apply(
                    lambda t: t.strftime("%m-%d %H:%M") if pd.notna(t) else "이번주 제출 없음")
                display_chk["제출횟수"] = checklist["제출횟수"].map(lambda n: f"{int(n)}회")

                # 표는 컬럼명이 라벨로 바뀐 뒤라, 라벨 기준으로 "양호" 응답을 찾을 수 있게 매핑한다.
                ok_by_label = {HW.CHECKLIST_LABELS[f]: HW.CHECKLIST_OK[f] for f in HW.CHECKLIST_FIELDS}

                def _highlight_chk(row):
                    return [
                        "background-color:#c81d25; color:white" if col in label_cols and _is_issue(row[col])
                        else _OK_CELL_STYLE if col in label_cols and str(row[col]).strip() == ok_by_label[col]
                        else "background-color:#fff3cd" if col == "제출횟수" and row[col] not in ("0회", "1회")
                        else ""
                        for col in row.index
                    ]

                st.dataframe(
                    display_chk.style.apply(_highlight_chk, axis=1),
                    width="stretch", hide_index=True,
                    column_config={"최근제출": st.column_config.TextColumn("최근 제출")},
                )
                st.caption("각 항목 = 이번주 그 지사의 최근 제출값(월요일 리셋). 빨간 칸 = 미흡·부족·신규 "
                           "발생 등 특이사항이 있는 항목 — \"특이사항\" 컬럼에 상세 내용이 기재됩니다. "
                           "초록 칸 = 정상(양호) 응답. "
                           "\"제출횟수\"가 2회 이상(노란 칸)이면 이번주 같은 지사에서 여러 번 제출된 것 — "
                           "폼에 응답자 구분 문항이 없어 표에는 가장 최근 제출값만 반영되니, 여러 명이 "
                           "제출했다면 서로 다른 값을 냈는지 지사 내에서 확인이 필요합니다.")

            photos = HW.weekly_photo_reports()
            st.markdown("---")
            st.subheader("📸 이번주 현장 활동 사진")
            st.link_button("📸 현장 사진 보고하기", _PHOTO_FORM_URL, width="stretch")
            st.caption("사진 업로드 문항이 있어 제출 시 구글 계정 로그인이 필요합니다. "
                       "지난주 이전에 올린 사진은 여기 목록과 인쇄 보고서에 나오지 않습니다(이번주 것만 반영).")

            # 구글폼(파일 업로드 문항이라 구글 로그인 필요)과 별개로, 여러 장을 한 번에
            # 등록하고 싶을 때 쓰는 두 번째 경로 — 엑셀에 지사·활동내용·사진파일명을 적고
            # 사진 파일들을 같이 올리면 파일명으로 매칭해 한 번에 등록한다. GITHUB_TOKEN이
            # secrets에 없으면(다른 배포본 등) 이 기능 자체를 숨긴다.
            if HW.github_write_enabled():
                with st.expander("📊 엑셀로 사진 일괄 등록 (여러 장 한 번에, 구글 로그인 불필요)"):
                    st.caption("① 아래 양식을 받아 지사·활동내용·사진파일명을 채우고 ② 그 엑셀과 "
                               "사진 파일들을 함께 올린 뒤 ③ 일괄 등록을 누르면, 파일명으로 자동 매칭해 "
                               "한 번에 등록됩니다.")
                    st.download_button(
                        "📥 엑셀 양식 다운로드", _bulk_photo_template_bytes(),
                        file_name="활동사진_일괄등록_양식.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                    # 등록에 성공하면 업로드 칸을 다시 빈칸으로 돌려서 "확실히 등록됐다"는 걸
                    # 눈으로 확인할 수 있게 한다. Streamlit은 위젯 key가 그대로면 값이 계속
                    # 남아있으므로, key 뒤에 세대(gen) 번호를 붙였다가 성공 시 번호를 올려서
                    # "새 위젯"으로 취급되게 만드는 방식을 쓴다. 결과 메시지는 rerun 직후에
                    # 바로 보여야 리셋된 빈 칸과 함께 한 화면에 뜬다.
                    _gen = st.session_state.setdefault("_bulk_uploader_gen", 0)
                    _pending = st.session_state.pop("_bulk_upload_result", None)
                    if _pending:
                        if _pending["success"]:
                            st.success(
                                f"✅ {len(_pending['success'])}건 등록 완료: " +
                                ", ".join(f"{s['branch']}({s['filename']})" for s in _pending["success"]) +
                                " — 화면에 반영되기까지 최대 1분 정도 걸릴 수 있습니다.")
                        if _pending["errors"]:
                            st.error("다음 행은 등록에 실패했습니다: " +
                                     " / ".join(f"{e['row']}행: {e['reason']}" for e in _pending["errors"]))

                    _xlsx_file = st.file_uploader("① 작성한 엑셀 파일", type=["xlsx"], key=f"_bulk_xlsx_{_gen}")
                    _photo_files = st.file_uploader(
                        "② 사진 파일들 (엑셀의 '사진파일명'과 이름이 정확히 같아야 매칭됩니다)",
                        type=["jpg", "jpeg", "png"], accept_multiple_files=True, key=f"_bulk_photos_{_gen}")
                    if st.button("📤 일괄 등록", width="stretch",
                                  disabled=not (_xlsx_file and _photo_files)):
                        try:
                            _bulk_df = pd.read_excel(_xlsx_file, sheet_name="사진목록")
                        except Exception:
                            try:
                                _bulk_df = pd.read_excel(_xlsx_file, sheet_name=0)
                            except Exception as _e:
                                st.error(f"엑셀을 읽지 못했습니다: {_e}")
                                _bulk_df = None
                        if _bulk_df is not None:
                            _missing_cols = {"지사", "활동내용", "사진파일명"} - set(_bulk_df.columns)
                            if _missing_cols:
                                st.error(f"엑셀에 필요한 칸이 없습니다: {', '.join(_missing_cols)} "
                                         "— 양식을 다시 받아서 칸 이름을 바꾸지 말고 채워주세요.")
                            else:
                                _file_map = {f.name: f.getvalue() for f in _photo_files}
                                _available_names = sorted(_file_map.keys())
                                _rows, _pre_errors = [], []
                                for _excel_row, (_, _r) in enumerate(_bulk_df.iterrows(), start=2):
                                    _branch = str(_r.get("지사") or "").strip()
                                    _fname = str(_r.get("사진파일명") or "").strip()
                                    # 옛 버전 양식(예시 행이 남아있던)을 계속 쓰는 사람이 있을 수
                                    # 있어, 지사·파일명 어느 쪽에 "예:"가 남아있어도 건너뛴다.
                                    if (not _branch or not _fname
                                            or _branch.startswith("예:") or _fname.startswith("예:")):
                                        continue
                                    _content = _file_map.get(_fname)
                                    if _content is None:
                                        # 여기서 바로 "업로드한 파일 목록"까지 보여줘야, 사용자가
                                        # 엑셀에 적은 이름과 실제 파일명이 어떻게 다른지 한눈에
                                        # 비교해서 고칠 수 있다(2026-08-06: 예시 행 "예: photo1.jpg"를
                                        # 그대로 남겨서 실패한 사고 이후 추가).
                                        _pre_errors.append({
                                            "row": _excel_row,
                                            "reason": (f"'{_fname}' 파일을 찾지 못했습니다 — 업로드한 "
                                                       f"사진 파일명: {', '.join(_available_names) if _available_names else '(없음)'}"),
                                        })
                                        continue
                                    _rows.append({
                                        "branch": _branch,
                                        "description": str(_r.get("활동내용") or "").strip(),
                                        "filename": _fname,
                                        "content": _content,
                                    })
                                if not _rows and not _pre_errors:
                                    st.warning("엑셀에서 등록할 행을 찾지 못했습니다 — 지사·사진파일명 "
                                               "칸에 실제 데이터를 입력했는지 확인해주세요.")
                                else:
                                    if _rows:
                                        with st.spinner(f"{len(_rows)}건 등록 중..."):
                                            _bulk_result = HW.submit_bulk_photos(_rows)
                                    else:
                                        _bulk_result = {"success": [], "errors": []}
                                    _bulk_result["errors"] = _pre_errors + _bulk_result["errors"]
                                    if _bulk_result["success"]:
                                        HW.load_bulk_photo_reports.clear()
                                        st.session_state["_bulk_uploader_gen"] += 1
                                    st.session_state["_bulk_upload_result"] = _bulk_result
                                    st.rerun()

            if not photos.empty:
                # 한 지사가 사진 여러 장을 한꺼번에 올리면 제출시각이 모두 같아 목록에서
                # 서로 구분되지 않는다 — 지사별로 번호를 매겨 캐러셀 캡션과 선택 목록에
                # 같은 번호를 쓰면, 캐러셀에서 본 사진을 번호로 바로 고를 수 있다.
                _photo_no = {}
                for _b, _g in photos.groupby("branch"):
                    for _i, _u in enumerate(
                            _g.sort_values("timestamp", ascending=False)["photo_url"].tolist(), start=1):
                        _photo_no[(_b, _u)] = _i

                # p.branch/p.photo_url은 구글시트 응답(외부 입력)이라 raw HTML(components.html)에
                # 꽂기 전에 반드시 이스케이프한다 — 폼이 드롭다운이라 지금은 안전해 보여도,
                # 시트를 직접 수정하거나 필드가 자유서술로 바뀌면 스크립트 삽입 경로가 된다.
                cards = "".join(
                    f'<div style="flex:0 0 auto; width:220px; scroll-snap-align:start;">'
                    f'<div style="position:relative;">'
                    f'<img src="{html.escape(p.photo_url)}" style="width:220px; height:220px; object-fit:cover; '
                    f'border-radius:8px; display:block;" />'
                    f'<div style="position:absolute; top:6px; left:6px; background:rgba(0,0,0,.72); color:white; '
                    f'font-size:13px; font-weight:700; padding:2px 8px; border-radius:10px;">'
                    f'{_photo_no.get((p.branch, p.photo_url), 0)}</div></div>'
                    f'<div style="font-size:12px; color:#333; background:rgba(255,255,255,.9); '
                    f'padding:4px 6px; border-radius:0 0 8px 8px;">'
                    f'{html.escape(str(p.branch))} '
                    f'#{_photo_no.get((p.branch, p.photo_url), 0)} · '
                    f'{p.timestamp.strftime("%m-%d %H:%M")}</div>'
                    f'</div>'
                    for p in photos.itertuples()
                )
                carousel_html = f"""
                <html><body style="margin:0;">
                <div style="position:relative; display:flex; align-items:center; font-family:sans-serif;">
                  <button onclick="document.getElementById('photoTrack').scrollBy({{left:-700,behavior:'smooth'}})"
                          style="border:none; background:#eee; border-radius:50%; width:32px; height:32px;
                                 cursor:pointer; flex-shrink:0; margin-right:6px; font-size:16px;">◀</button>
                  <div id="photoTrack" style="display:flex; gap:12px; overflow-x:auto; scroll-snap-type:x mandatory;
                       scroll-behavior:smooth; padding:4px;">
                    {cards}
                  </div>
                  <button onclick="document.getElementById('photoTrack').scrollBy({{left:700,behavior:'smooth'}})"
                          style="border:none; background:#eee; border-radius:50%; width:32px; height:32px;
                                 cursor:pointer; flex-shrink:0; margin-left:6px; font-size:16px;">▶</button>
                </div>
                </body></html>
                """
                components.html(carousel_html, height=270, scrolling=False)
                st.caption("사진 아래 캡션 = 올린 지사 · 제출시각. 좌우 화살표 또는 마우스 드래그/트랙패드로 넘길 수 있습니다.")

                # 인쇄 보고서에 넣을 사진을 지사별로 직접 고른다 — 기본값은 최신 사진이라
                # 그대로 둬도 되고, 보고에 더 적합한 장면이 있으면 바꿀 수 있게 한다.
                # 확인 버튼을 누르면 Streamlit이 재실행되면서 expander가 다시 접혀 미리보기가
                # 안 보인다 — 확인 상태에서는 펼친 채로 유지한다.
                with st.expander("🖨️ 보고서에 넣을 사진 고르기 (기본: 지사별 최신 사진)",
                                  expanded=bool(st.session_state.get("_photo_pick_confirmed"))):
                    picks = dict(st.session_state.get("_report_photo_pick", {}))
                    changed = False
                    # 사진이 있는 지사만 드롭다운에 나오면 "빠진 지사"가 안 보여
                    # 헷갈리므로, 12개 지사를 전부 늘어놓고 없는 지사는 선택 불가한
                    # "금주 활동 사진 없음" 한 줄로 표시한다.
                    for _b in HW.branch_order():
                        _bp = photos[photos["branch"] == _b].sort_values("timestamp", ascending=False)
                        if _bp.empty:
                            st.selectbox(_b, ["금주 활동 사진 없음"], index=0, disabled=True,
                                         key=f"_photopick_{_b}")
                            continue
                        _labels = ["최신 사진 (기본)"] + [
                            f'#{_photo_no.get((_b, r.photo_url), 0)} · {r.timestamp.strftime("%m-%d %H:%M")}'
                            for r in _bp.itertuples()]
                        _urls = [None] + _bp["photo_url"].tolist()
                        _cur = picks.get(_b)
                        _idx = _urls.index(_cur) if _cur in _urls else 0
                        _sel = st.selectbox(f"{_b} ({len(_bp)}장)", _labels, index=_idx,
                                            key=f"_photopick_{_b}")
                        _new = _urls[_labels.index(_sel)]
                        if _new != picks.get(_b):
                            changed = True
                            if _new is None:
                                picks.pop(_b, None)
                            else:
                                picks[_b] = _new
                    if changed:
                        st.session_state["_report_photo_pick"] = picks

                    # 고른 사진이 실제로 반영됐는지 눈으로 확인할 수 있게, 확인 버튼을 누르면
                    # 보고서에 들어갈 사진을 지사별로 미리 보여준다(선택 안 한 지사는 최신 사진).
                    if st.button("✅ 선택 확인 (보고서에 들어갈 사진 미리보기)", width="stretch"):
                        st.session_state["_photo_pick_confirmed"] = True
                    if st.session_state.get("_photo_pick_confirmed"):
                        _picked = st.session_state.get("_report_photo_pick", {})
                        _branches = HW.branch_order()
                        _with_photo = [b for b in _branches if b in set(photos["branch"])]
                        st.success(f"보고서에 들어갈 사진 — 전체 {len(_branches)}개 지사 중 "
                                   f"{len(_with_photo)}곳 사진 있음(직접 고른 지사 {len(_picked)}곳, "
                                   f"나머지는 최신 사진), {len(_branches) - len(_with_photo)}곳은 "
                                   "사진 없음으로 인쇄됩니다.")
                        for _row in [_branches[i:i + 4] for i in range(0, len(_branches), 4)]:
                            for _col, _b in zip(st.columns(4), _row):
                                _bp = photos[photos["branch"] == _b].sort_values(
                                    "timestamp", ascending=False)
                                with _col:
                                    if _bp.empty:
                                        st.info(f"{_b} · 금주 활동 사진 없음")
                                        continue
                                    _u = _picked.get(_b) or _bp.iloc[0]["photo_url"]
                                    _n = _photo_no.get((_b, _u), 0)
                                    _manual = _b in _picked
                                    st.image(_u, width="stretch")
                                    st.caption(f'{_b} #{_n} · {"직접 선택" if _manual else "최신(기본)"}')
                    st.caption("목록의 번호(#1, #2…)는 위 사진 왼쪽 위에 붙은 번호와 같습니다 — 같은 시각에 "
                               "여러 장을 올렸어도 번호로 구분해 고를 수 있습니다. "
                               "고른 사진은 위 \"🖨️ 주간 보고서 인쇄\" 버튼으로 뽑는 보고서에 반영됩니다. "
                               "선택은 이 브라우저 세션에서만 유지되며, 새로고침하면 최신 사진으로 돌아갑니다.")
            else:
                st.info("아직 업로드된 사진이 없습니다.")

            st.markdown("---")
            with st.expander("⚖️ 고용노동부 규정 개정 참고사항"):
                st.markdown(
                    "**'세방(주) 온열질환 예방지침' REV2(사내 기준)**: 주의 33℃ · 경고 35℃ · 위험 38℃\n\n"
                    "2025년 7월 17일 「산업안전보건기준에 관한 규칙」 개정으로 **체감온도 31℃ 이상을 "
                    "'폭염 작업'으로 정의**하고, 그 구간에서 체감온도 측정·기록과 조치(냉방·통풍장치/"
                    "작업시간 조정/휴식 중 1개 이상) 이행이 법적 의무가 됐습니다."
                )
                st.caption("출처: 고용노동부 정책자료(2025.7), 김·장 법률사무소 규정 해설 등 — 개정 진행 "
                           "경과가 있어 법령 원문(moel.go.kr) 확인을 권합니다.")

            with st.expander("🚑 온열질환 응급처치 빠른 참조"):
                st.markdown("""
| 질환 | 주요 증상 | 응급처치 |
|---|---|---|
| **열사병**(응급, 즉시 119) | 의식저하·헛소리, 체온 40℃↑, 피부가 뜨겁고 건조함 | 즉시 119 신고 → 그늘/에어컨 공간으로 이동 → 옷을 벗기고 물을 뿌리며 부채질 등 적극적으로 체온을 낮춤 → **의식이 없으면 음료를 먹이지 않음** |
| **열탈진** | 어지럼증·두통·메스꺼움, 많은 땀 | 그늘로 이동, 옷을 헐렁하게, 시원한 물을 조금씩 섭취, 30분 내 호전 없으면 병원 이송 |
| **열경련** | 팔다리·복부 근육 경련 | 그늘에서 휴식, 이온음료나 물 섭취, 경련 부위 스트레칭 |
| **열실신** | 어지럼증과 함께 순간적인 의식 소실 | 시원한 곳에 눕히고 다리를 심장보다 높게 올림 |
""")
                st.caption("출처: 질병관리청·고용노동부 공개 자료 기준 일반 응급처치 요령. 실제 상황에서는 항상 119 신고를 우선하세요.")

            st.markdown("---")
            st.subheader("📈 주차별 최고 체감온도 추이 (지사별)")
            wmax = HW.weekly_max_by_branch(obs)
            wmax["week_str"] = wmax["week"].dt.strftime("%Y-%m-%d")
            wmax_chart = wmax.dropna(subset=["apparent_temp"]).copy()
            wmax_chart["단계"] = wmax_chart["apparent_temp"].apply(HW.temp_level).fillna("정상")
            wmax_chart["라벨"] = wmax_chart["apparent_temp"].map(lambda t: f"{t:.1f}°")
            level_colors = {"정상": HW.NORMAL_COLOR, **HW.LEVEL_COLOR}
            fig_temp = px.bar(wmax_chart, x="branch", y="apparent_temp", color="단계", facet_col="주차", height=380,
                               category_orders={"단계": ["정상", *HW.LEVEL_ORDER], "branch": HW.branch_order(),
                                                 "주차": ["지난주", "이번주"]},
                               color_discrete_map=level_colors, text="라벨")
            fig_temp.update_traces(textposition="outside", textfont=dict(size=13, color="white"), cliponaxis=False)
            for lvl, y in [("주의", 33), ("경고", 35), ("위험", 38)]:
                fig_temp.add_hline(y=y, line_dash="dot", line_color=HW.LEVEL_COLOR[lvl],
                                    annotation_text=lvl, annotation_position="right")
            fig_temp.update_layout(margin=dict(l=0, r=0, t=40, b=0), yaxis_title="체감온도(°C)", legend_title="단계")
            # 주차 라벨 옆에 실제 날짜(그 주 월요일)도 같이 보여준다.
            week_dates = dict(zip(wmax["주차"], wmax["week_str"]))
            fig_temp.for_each_annotation(
                lambda a: a.update(text=f"{a.text} ({week_dates.get(a.text, '')}~)") if a.text in week_dates else a)
            fig_temp.update_xaxes(title="")
            # range만 지정하면 데스크톱에선 문제없지만, 모바일에서 반응형 리사이즈가
            # 일어날 때 autorange가 다시 켜지면서 막대(바)형 축 특성상 0부터 다시
            # 그려지는 경우가 있어(확인됨) autorange=False를 명시적으로 고정한다.
            fig_temp.update_yaxes(range=[25, 40], autorange=False, rangemode="normal")
            st.plotly_chart(fig_temp, width="stretch")
            st.caption("왼쪽 = 지난주, 오른쪽 = 이번주(진행 중). 각 지사·주차의 최고 체감온도만 표시합니다.")

            st.markdown("---")
            st.subheader("🌡️ 지사별 폭염 단계 발생 현황(주차별)")
            weekly = HW.weekly_alert_counts_by_branch(notif)
            fig_alert = px.bar(weekly, x="branch", y="발령횟수", color="level", barmode="stack",
                                facet_col="주차", height=380,
                                category_orders={"level": HW.LEVEL_ORDER, "branch": HW.branch_order(),
                                                  "주차": ["지난주", "이번주"]},
                                color_discrete_map=HW.LEVEL_COLOR)
            fig_alert.update_layout(margin=dict(l=0, r=0, t=40, b=0), yaxis_title="발령 건수", legend_title="단계")
            fig_alert.update_xaxes(title="")
            st.plotly_chart(fig_alert, width="stretch")
            st.caption("왼쪽 = 지난주, 오른쪽 = 이번주(진행 중 — 아직 발령이 없으면 빈 패널로 표시됩니다).")

            if not notif.empty:
                with st.expander("📋 최근 발령 이력 전체"):
                    st.dataframe(
                        notif.sort_values("sent_at", ascending=False)[["branch", "site", "level", "apparent_temp", "sent_at", "status"]],
                        width="stretch", hide_index=True)

                    st.markdown("---")
                    st.markdown("##### 🚀 지금 발송 (파일럿 · 수동)")
                    st.caption("자동 발송은 아직 꺼져 있습니다 — 담당자가 버튼을 직접 눌러야만 실제 "
                               "문자가 나갑니다. 시연·검증이 끝나면 자동 발송으로 전환할 수 있습니다.")

                    _pilot_ready = HW.solapi_configured()
                    _send_branches = HW.branch_order() or HW.branches_with_recipients()
                    _send_branch = st.selectbox("지사 선택", _send_branches, key="_send_branch")
                    _recipients = HW.get_recipients(_send_branch)
                    _demo_mode = not (_pilot_ready and _recipients)

                    if _demo_mode:
                        _why = ("Solapi 연동 정보(API Key/Secret/발신번호)가 secrets에 아직 없어"
                                 if not _pilot_ready else f"{_send_branch} 지사에 담당자가 아직 등록되지 않아")
                        st.info(f"🧪 시연 모드 — {_why} 실제 문자는 나가지 않고 화면 흐름만 보여줍니다.")
                    _display_recipients = _recipients if _recipients else [HW.DEMO_RECIPIENT]

                    st.caption(f"수신자 {len(_display_recipients)}명: " +
                               ", ".join(f"{r['name']}({r['phone']})" for r in _display_recipients))

                    _branch_notif = notif[notif["branch"] == _send_branch].sort_values(
                        "sent_at", ascending=False)
                    if _branch_notif.empty:
                        _default_msg = f"[{_send_branch}] 온열질환 예방 안내 (파일럿 시연용)"
                    else:
                        _latest = _branch_notif.iloc[0]
                        _default_msg = HW.build_alert_message(
                            _send_branch, _latest["level"], _latest["apparent_temp"], _latest["sent_at"])

                    _msg = st.text_area("발송 내용 (필요하면 수정 가능)", value=_default_msg,
                                         height=110, key=f"_send_msg_{_send_branch}")

                    _button_label = (f"🧪 {_send_branch} 지사 발송 시연 (실제 문자 안 나감)" if _demo_mode else
                                      f"🚀 {_send_branch} 지사 담당자 {len(_display_recipients)}명에게 지금 발송")
                    if st.button(_button_label, type="primary", width="stretch"):
                        if _demo_mode:
                            st.success("✅ (시연) 아래 담당자에게 발송되는 흐름입니다 — 실제로 전송되지는 "
                                       "않았습니다: " + ", ".join(
                                           f"{r['name']}({r['phone']})" for r in _display_recipients))
                        else:
                            with st.spinner("발송 중..."):
                                _send_result = HW.send_alert_sms(_send_branch, _msg)
                            if _send_result["sent"]:
                                st.success("✅ 발송 완료: " + ", ".join(
                                    f"{s['name']}({s['phone']})" for s in _send_result["sent"]))
                            if _send_result["failed"]:
                                st.error("❌ 발송 실패: " + " / ".join(
                                    f"{f['name']}: {f['reason']}" for f in _send_result["failed"]))


# ================================================================== TAB 일일 안전보건 브리핑
# 2026-08-13 신설(자연재해 통합경보 탭 대체 — 배포 환경 동기화 문제가 반복돼
# 사용자가 그 탭을 없애고 이걸로 바꿔달라고 요청). "온도를 조져보자/
# daily_report.py"가 매일 만드는 뉴스+DART+판결문+날씨 통합 브리핑 이메일과
# 완전히 같은 내용을 그대로 보여준다 — 대시보드가 직접 API/LLM을 호출하지
# 않고 미리 만들어진 스냅샷(heatwave_data/daily_briefing_latest.html)만
# 읽는다(비용 통제, 사용자 지적 반영). 뉴스/판결문 원문 등 대외비 성격의
# 원자료를 포함하고 있어, 사용자 요청(2026-08-13)에 따라 전사현황/지사상세/
# 점검편향분석(tab1~3)과 같은 2차 비밀번호(ANALYSIS_PASSWORD) 게이트를 건다.
with tab_briefing:
    # 2026-08-13 사용자 요청 — 전사현황/지사상세/점검편향분석(tab1~3)과 같은
    # 2차 비밀번호(ANALYSIS_PASSWORD)로 잠근다. _analysis_gate()가 잠겨 있으면
    # 입력창을 그리고 False를 돌려주므로, 그 경우 아래 실제 내용은 그리지 않는다.
    if _analysis_gate("tab_briefing"):
        st.subheader("📰 일일 안전보건 브리핑")
        st.caption("뉴스·DART 공시·판결문·날씨 특보를 종합한 안전보건관리책임자용 일일 브리핑입니다. "
                   "매일 자동 실행되는 온도를 조져보자/daily_report.py가 만든 최신 결과를 그대로 보여줍니다.")
        _briefing_html, _briefing_mtime = HW.load_daily_briefing_html()
        if not _briefing_html:
            st.info("아직 생성된 브리핑이 없습니다. `온도를 조져보자` 프로젝트에서 daily_report.py가 "
                    "한 번 이상 실행된 뒤 sync_heatwave_data.py로 동기화되면 여기 표시됩니다.")
        else:
            if _briefing_mtime is not None:
                st.caption(f"🕒 이 스냅샷 동기화 시각: {_briefing_mtime.strftime('%Y-%m-%d %H:%M')} "
                           "(브리핑 자체의 수집 기준 시각은 아래 본문 상단에 별도로 표시됩니다)")
            components.html(_briefing_html, height=1400, scrolling=True)


# ------------------------------------------------------- 분석 탭 2차 비밀번호 게이트
# 폭염 대응(tab4) 블록을 분석 탭보다 먼저 그린다 — 잠겨 있을 때 st.stop()으로 이후
# 실행을 멈춰도 운영 화면인 폭염 탭은 이미 렌더링돼 정상 동작하게 하기 위함이다.
# (탭이 화면에 보이는 순서는 st.tabs() 선언 순서라, 코드 순서를 바꿔도 UI는 그대로다.)
if not _analysis_unlocked():
    with tab1:
        _analysis_gate("tab1")
    with tab2:
        _analysis_gate("tab2")
    with tab3:
        _analysis_gate("tab3")
    st.stop()


# ================================================================== TAB1 전사 현황
with tab1:
    st.success(
        "🧭 **이 지수를 어떻게 볼까요?** — 위험지수는 지사를 평가·처벌하는 점수가 **아닙니다.** "
        "점검을 활발하고 꼼꼼하게 한 지사일수록 위험을 많이 발굴해 지수가 높게 나올 수 있으며, "
        "이는 **안전관리 활동이 적정하게 이뤄지고 있다는 긍정적 신호**로 해석합니다. "
        "본사는 이 값을 자원 배분(교육·예산·인력)의 우선순위 참고자료로만 사용합니다.",
        icon="🧭")

    if SEV_COL == "LLM심각도":
        st.info(
            "🧪 **LLM 하이브리드 모드** — 지적내용의 문맥(작업 위치·상황)까지 반영한 파일럿 결과입니다. "
            "규칙기반 대비 당진지사가 상위로, 부산지사가 하위로 순위가 조정됩니다.", icon="🧪")

    c1, c2 = st.columns([1, 1])
    with c1:
        NORM = {"총량": "위험지수", "사업장당": "사업장당지수", "안전관리자당": "관리자당지수"}
        basis_label = st.radio("보정 기준", list(NORM.keys()), horizontal=True,
                               help="지사마다 사업장·안전관리자 수가 달라, 규모로 나눈 값도 함께 볼 수 있습니다.")
        basis = NORM[basis_label]
        ri_sorted = ri.sort_values(basis, ascending=True)
        fig = px.bar(ri_sorted, x=basis, y="지사", orientation="h",
                     text=basis, height=430,
                     hover_data=["지적건수", "평균심각도", "안전관리자수", "사업장수"])
        fig.update_traces(marker_color=CHART_ACCENT, textposition="outside", cliponaxis=False)
        fig.update_layout(margin=dict(l=0, r=30, t=30, b=0),
                          xaxis_title="", yaxis_title="",
                          title=f"지사별 위험·점검활동 지수 ({basis_label})")
        st.plotly_chart(fig, width="stretch")
        st.caption("높다고 위험한 지사가 아니라, 위험을 많이 발굴했거나 규모가 큰 지사일 수 있습니다. "
                   "오른쪽 '양 vs 질'로 나눠서 보세요.")
    with c2:
        st.markdown("**점검의 '양'(건수) vs '질'(평균 심각도)**")
        fig = px.scatter(ri, x="지적건수", y="평균심각도", size="가중점수",
                         color="평균심각도", color_continuous_scale="Blues",
                         text="지사", height=430, size_max=45)
        fig.update_traces(textposition="top center")
        fig.add_vline(x=ri["지적건수"].median(), line_dash="dot", line_color="gray")
        fig.add_hline(y=ri["평균심각도"].median(), line_dash="dot", line_color="gray")
        fig.update_layout(coloraxis_showscale=False, margin=dict(l=0, r=0, t=10, b=0),
                          xaxis_title="점검 건수(활동량)", yaxis_title="평균 심각도(위험 정도)")
        st.plotly_chart(fig, width="stretch")
        st.caption("오른쪽 = 점검 활발(바람직) · 위쪽 = 심각한 지적 비중 높음. "
                   "'점검을 많이 한 것'과 '위험이 심각한 것'은 다릅니다.")

    with st.expander("📋 지사별 상세 표 (건수·규모·지수)"):
        st.dataframe(
            ri[["지사", "지적건수", "평균심각도", "안전관리자수", "사업장수",
                "위험지수", "사업장당지수", "관리자당지수"]],
            width="stretch", hide_index=True)

# ================================================================== TAB2 지사 상세
# 워드클라우드 불용어 — 문법 조각 · 점검 '과정' 메타어 · 지명
_STOP = ("있 없 및 등 위 시 후 전 중 내 외 관련 대한 하여 위해 인한 인해 따른 통해 실시 확인 조치 필요 요함 요망 "
         "발생 상태 작업 현장 근로자 사용 부분 경우 이동 진행 완료 요청 가능 대해 으로 에서 하는 되는 방지 미흡 "
         "이나 또는 지사 내용 사항 여부 당사 소속 운전 원인 결과 예정 이상 우측 좌측 일부 해당 예방 조치요망 "
         "위험 안전 관리 오늘 하기 되어 있는 없는 통한 대비 이후 이전 위한 우려 존재 가능성 있음 없음 등의 등을 "
         "등이 대하여 관하여 인하여 위하여 하도록 하며 하고 이고 이며 되며 따라 아래 다음 지적 코칭 개선 점검 "
         "확인함 실시함 조치함 요청함 점검일 일지 관리자 담당 지금 현문 세트 서류 인원 주변 사무동 게시 위치 "
         "당시 부착 재해 저하 상시 수시 조치예정 조치완료 미실시 "
         "강원 경남 경북 경인 광양 당진 목포 부산 삼천포 전북 울산 본사 인천 경기 포항 군산 동해 창원")
STOPWORDS = set(_STOP.split())


def tokenize(texts):
    cnt = {}
    for t in texts:
        for w in re.findall(r"[가-힣]{2,}", str(t)):
            if w in STOPWORDS or w.endswith("지사"):
                continue
            cnt[w] = cnt.get(w, 0) + 1
    return cnt


def growth_wordcloud(df_branch, col, max_words=30):
    """급증 위험요인=빨강. 최근 60일 vs 이전 60일 빈도 증가율로 색상. 단어 수 제한으로 식별성 확보."""
    end_ = df_branch["점검일자"].max()
    recent = df_branch[df_branch["점검일자"] >= end_ - pd.Timedelta(days=60)]
    prior = df_branch[(df_branch["점검일자"] < end_ - pd.Timedelta(days=60)) &
                      (df_branch["점검일자"] >= end_ - pd.Timedelta(days=120))]
    fr, fp, total = tokenize(recent[col]), tokenize(prior[col]), tokenize(df_branch[col])
    if not total:
        return None
    growth = {w: (fr.get(w, 0) - fp.get(w, 0)) / (fp.get(w, 0) + 1) for w in total}

    def color_func(word, **kw):
        g = growth.get(word, 0)
        if g >= 1.0:   return "#d62728"
        if g > 0:      return "#e8862c"
        return "#7f8fa6"

    wc = WordCloud(font_path=FONT_PATH, background_color="white",
                   width=720, height=420, max_words=max_words, color_func=color_func,
                   prefer_horizontal=0.95, margin=6).generate_from_frequencies(total)
    fig, ax = plt.subplots(figsize=(8, 4.6))
    ax.imshow(wc, interpolation="bilinear"); ax.axis("off")
    fig.tight_layout(pad=0)
    return fig


with tab2:
    sel = st.selectbox("지사 선택", ri["지사"].tolist())
    dfb = insp[insp["지사"] == sel]
    rrow = ri[ri["지사"] == sel].iloc[0]

    cA, cB = st.columns([1, 2])
    with cA:
        val = float(rrow["위험지수"])
        gfig = go.Figure(go.Indicator(
            mode="gauge+number", value=val, title={"text": f"{sel} 위험·활동 지수"},
            gauge={"axis": {"range": [0, 100]},
                   "bar": {"color": CHART_ACCENT},
                   "steps": [{"range": [0, 50], "color": "#F2F3F3"},
                             {"range": [50, 80], "color": SEBANG_LIGHT_GRAY_1},
                             {"range": [80, 100], "color": SEBANG_LIGHT_GRAY_2}]}))
        gfig.update_layout(height=260, margin=dict(l=10, r=10, t=50, b=10))
        st.plotly_chart(gfig, width="stretch")
        st.caption(f"점검 {int(rrow['지적건수'])}건 · 평균 심각도 {rrow['평균심각도']:.2f} · "
                   f"사업장 {int(rrow['사업장수'])}곳 · 안전관리자 {int(rrow['안전관리자수'])}명")
        st.markdown("**위험분류별 지적 분포**")
        vc = dfb[dfb["위험분류"].isin(S.PHYSICAL_HAZARDS)]["위험분류"].value_counts()
        st.bar_chart(vc, height=240, color=CHART_ACCENT)
    with cB:
        st.markdown("**동적 워드클라우드** — 🔴급증 · 🟠증가 · ⚪안정 (최근 60일 대비) · 상위 30개 단어")
        w1, w2 = st.columns(2)
        with w1:
            st.markdown("🔎 **점검(코칭)내용** — 무엇을 지적했나")
            fig = growth_wordcloud(dfb, "지적내용_평문")
            if fig is not None:
                st.pyplot(fig)
            else:
                st.info("텍스트 부족")
        with w2:
            st.markdown("🛠️ **조치내용** — 어떻게 개선했나")
            fig = growth_wordcloud(dfb, "조치내용_평문")
            if fig is not None:
                st.pyplot(fig)
            else:
                st.info("텍스트 부족")

    st.markdown("---")
    cc1, cc2 = st.columns(2)
    with cc1:
        st.subheader("🎯 점검 사각지대")
        st.caption("과거 인적재해가 있었으나 최근 점검에서 다루지 않은 위험분류 (질병성 제외)")
        bs = S.blind_spots(insp, acc, exclude_disease=True)
        bsb = bs[bs["지사"] == sel] if not bs.empty else bs
        if bsb is None or bsb.empty:
            st.success("최근 점검에서 놓친 과거재해 위험분류가 없습니다.")
        else:
            st.dataframe(bsb[["위험분류", "과거재해", "최근점검", "상태"]],
                         width="stretch", hide_index=True, height=220)
    with cc2:
        st.subheader("📜 과거 인적재해 이력")
        pa = acc[(acc["지사"] == sel) & (acc["재해성격"] == "인적재해")]
        if pa.empty:
            st.info("인적재해 이력 없음")
        else:
            st.caption(f"총 {len(pa)}건 (산재 {int((pa['산재구분']=='산재').sum())} · "
                       f"공상 {int((pa['산재구분']=='공상').sum())})")
            hist = pa.groupby("위험분류").size().sort_values(ascending=False).reset_index(name="건수")
            st.dataframe(hist, width="stretch", hide_index=True, height=220)

    # ---- 분류 재검토 제안 (AI 텍스트 분석) ----
    st.markdown("---")
    st.subheader("🔧 위험분류 재검토 제안 (AI 텍스트 분석)")
    st.caption("LLM이 지적내용을 읽고, 규칙기반으로 붙은 위험분류가 실제 내용과 맞지 않아 보이는 사례를 찾은 결과입니다. "
               "같은 유형(예: 안전모 미착용)이 지사마다 다르게 분류되는 문제를 발견할 수 있습니다.")
    mis_all = insp[insp["위험분류_재검토"].notna()].copy()
    if not mis_all.empty:
        mis_all["지적내용"] = mis_all["지적내용_평문"].astype(str).str.replace(r"\([^)]*\)", "", regex=True)
    mis_b = mis_all[mis_all["지사"] == sel] if not mis_all.empty else mis_all
    if mis_b is None or mis_b.empty:
        st.info(f"{sel}은 현재 재검토 후보가 없습니다. (아래 전사 목록에서 다른 지사 사례를 볼 수 있습니다)")
    else:
        st.write(f"**{sel}** 재검토 후보 {len(mis_b)}건")
        st.dataframe(
            mis_b[["위험분류", "위험분류_재검토", "지적내용"]].rename(
                columns={"위험분류": "원본 분류", "위험분류_재검토": "AI 재검토 제안"}),
            width="stretch", hide_index=True)
    if not mis_all.empty:
        with st.expander(f"🔎 전사 재검토 후보 전체 {len(mis_all)}건 — 지사 간 분류 일관성 점검"):
            st.dataframe(
                mis_all[["지사", "위험분류", "위험분류_재검토", "지적내용"]].rename(
                    columns={"위험분류": "원본 분류", "위험분류_재검토": "AI 재검토 제안"}),
                width="stretch", hide_index=True)

# ================================================================== TAB3 점검 편향분석
with tab3:
    st.info(
        "이 지사(안전관리자)의 점검이 **특정 위험에만 쏠려 있지 않은지**, 그리고 **실제 사고가 나는 유형을 "
        "잘 점검하고 있는지**를 자체 점검하는 화면입니다. 잘잘못을 가리는 것이 아니라 사각지대를 스스로 "
        "발견하기 위한 참고 지표입니다.", icon="🔍")
    st.caption("※ 근골격 등 질병성(지연발현) 재해는 현장 실시간 점검으로 포착이 어려워 편향분석에서 제외합니다.")

    st.subheader("점검자별 위험분류 다양성 지수")
    st.caption("지수가 낮을수록 특정 위험분류에 편중된 점검입니다. 막대를 클릭하거나 아래에서 점검자를 선택하면 상세가 열립니다.")
    bias = S.inspector_bias(insp)
    figb = px.bar(bias, x="다양성지수", y="점검자", orientation="h", color="다양성지수",
                  color_continuous_scale="RdYlGn", range_color=[0.6, 1.0],
                  hover_data=["지사", "점검건수", "최다위험", "최다비중", "다룬위험종류"], height=420)
    figb.update_layout(yaxis={"categoryorder": "total descending"},
                       coloraxis_showscale=False, margin=dict(l=0, r=0, t=10, b=0))
    ev = st.plotly_chart(figb, width="stretch", on_select="rerun", key="bias_chart")

    picked = None
    try:
        pts = ev.selection.points if ev and ev.selection else []
        if pts:
            picked = pts[0].get("y")
    except Exception:
        picked = None
    names = bias["점검자"].tolist()
    idx = names.index(picked) if picked in names else 0
    who = st.selectbox("점검자 상세 보기", names, index=idx, key="bias_pick")

    det = S.inspector_detail(insp, who)
    brow = bias[bias["점검자"] == who].iloc[0]
    st.markdown(f"#### 👷 {who} · {det['지사']} · 점검 {det['점검건수']}건 · "
                f"다양성지수 {brow['다양성지수']} (최다: {brow['최다위험']} {brow['최다비중']:.0f}%)")
    dc1, dc2 = st.columns(2)
    with dc1:
        st.markdown("**어떤 위험을 점검했나** (위험분류 분포)")
        st.bar_chart(det["위험분류분포"], color=CHART_ACCENT, height=240)
    with dc2:
        st.markdown("**어느 사업장에서 점검했나** (사업장 분포)")
        st.bar_chart(det["사업장분포"], color=CHART_ACCENT, height=240)
    if det["소외위험"]:
        st.warning(f"🔎 **{who}**님이 최근 한 번도 다루지 않은 물리위험: **{', '.join(det['소외위험'])}** "
                   "→ 다음 점검 시 의식적으로 확인 권장")
    else:
        st.success(f"{who}님은 주요 물리위험을 고르게 점검하고 있습니다.")

    st.markdown("---")
    st.subheader("지사별 갭분석 — 점검 vs 실제 사고")
    st.caption("선택한 지사의 위험분류별 [점검 비중] vs [사고 비중]. 사고비중이 점검비중보다 크면 "
               "그 지사의 점검이 실제 위험을 놓치고 있을 수 있습니다. (질병성 제외)")
    gsel = st.selectbox("지사 선택", INSP_BRANCHES,
                        index=INSP_BRANCHES.index(sel) if sel in INSP_BRANCHES else 0, key="gap_branch")
    gb = S.coverage_gap_by_branch(insp, acc, gsel, exclude_disease=True)
    if gb.empty or gb[["점검비중", "사고비중"]].to_numpy().sum() == 0:
        st.info(f"{gsel}: 비교할 사고 이력 또는 점검 데이터가 부족합니다.")
    else:
        gm = gb.reset_index().rename(columns={"index": "위험분류"}).melt(
            id_vars="위험분류", value_vars=["점검비중", "사고비중"],
            var_name="구분", value_name="비중")
        figg = px.bar(gm, x="위험분류", y="비중", color="구분", barmode="group", height=380,
                      color_discrete_map={"점검비중": SEBANG_GREEN, "사고비중": "#e45756"})
        figg.update_layout(margin=dict(l=0, r=0, t=10, b=0), legend_title="", yaxis_title="비중(%)")
        st.plotly_chart(figg, width="stretch")
        over = gb[gb["갭(사고-점검)"] > 5]
        if not over.empty:
            st.warning("⚠️ **" + gsel + "**에서 사고 대비 점검이 부족한 위험분류: "
                       + ", ".join(f"**{h}**(+{gb.loc[h,'갭(사고-점검)']:.0f}%p)" for h in over.index))

# ================================================================== TAB4 폭염 대응
